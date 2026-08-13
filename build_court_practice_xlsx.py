#!/usr/bin/env python3
"""Excel по судебной практике — тот же формат, что сводки по законодательству."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from build_seven_countries_pack import COUNTRIES

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
LINK_FONT = Font(color="0563C1", name="Calibri", size=10)
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79", name="Calibri")
LABEL_FONT = Font(bold=True, name="Calibri", size=11)

ROOT_EDO = Path("Анализ_судебной_практики_ЭДО.xlsx")
ROOT_OT = Path("Анализ_судебной_практики_ЭДО_в_ОТ.xlsx")
ROOT_CMP = Path("Сравнение_судебных_практик.xlsx")

RF_EDO = [
    "Российская Федерация (ориентир)",
    "ВС РФ / арбитражные и СОЮ",
    "Постановление Пленума ВС РФ от 23.06.2015 № 25 (электронная переписка); ст. 75 АПК, ст. 71 ГПК; 63‑ФЗ",
    "прямая (линия практики, не одно дело)",
    "Электронные документы и переписка принимаются, если можно установить достоверность; скриншоты оцениваются свободно; УКЭП даёт максимальную силу",
    "Электронный документ не отвергается только из‑за формы, но нет статутной презумпции как LEC 326.4 и нет фильтра 106B",
    "Сравнивать с зарубежными делами нужно по допустимости и виду ЭП, а не по «журналу инструктажа» — его РФ вывела законом (ст. 22.1(3) ТК), а не судом",
    "https://www.consultant.ru/document/cons_doc_LAW_181602/",
]

RF_OT = [
    "Российская Федерация (ориентир)",
    "Законодатель, не суд",
    "ТК РФ ст. 22.1 ч. 3; ПП 2464; практика бумажных журналов инструктажа",
    "статут (не судебный запрет e‑журнала)",
    "Журналы инструктажа и акты НС исключены из кадрового ЭДО. Суды по ОТ обычно проверяют факт инструктажа по бумажному журналу",
    "Запрет носителя — законодательный выбор, среди зарубежных карточек нет судебного правила «электронный журнал ОТ всегда ничтожен»",
    "Для гл. 3 НИР: зарубежная практика спорит о содержании обучения и силе подписи; РФ закрыла вопрос носителя нормой ТК",
    "https://www.consultant.ru/document/cons_doc_LAW_34683/",
]


def style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN


def write_row(ws, r: int, values: list, height: int = 78) -> None:
    for c, val in enumerate(values, start=1):
        cell = ws.cell(r, c, val)
        cell.alignment = WRAP
        cell.border = THIN
        cell.font = BODY_FONT
        text = str(val)
        if text.startswith("http://") or text.startswith("https://"):
            cell.hyperlink = text
            cell.font = LINK_FONT
    ws.row_dimensions[r].height = height


def set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def case_url(case: dict) -> str:
    srcs = case.get("sources") or []
    if srcs:
        return srcs[0][1]
    return ""


def meaning_text(case: dict) -> str:
    items = case.get("meaning") or []
    return " ".join(items[:2]) if items else ""


def analog_label(case: dict) -> str:
    return "аналогия" if case.get("analog") else "прямая"


def edo_row(c: dict) -> list:
    case = c["case_edo"]
    return [
        c["name"],
        case["court"],
        case["cite"],
        analog_label(case),
        case["title"],
        case["holding"],
        meaning_text(case),
        case_url(case),
    ]


def ot_row(c: dict) -> list:
    case = c["case_ot"]
    return [
        c["name"],
        case["court"],
        case["cite"],
        analog_label(case),
        case["title"],
        case["holding"],
        meaning_text(case),
        case_url(case),
    ]


def add_summary_sheet(wb, title: str, subtitle: str, rows: list[tuple[str, str]]) -> None:
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = subtitle
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 40
    ws["A4"] = "Блок"
    ws["B4"] = "Содержание"
    style_header_row(ws, 4, 2)
    for i, (a, b) in enumerate(rows, start=5):
        ws.cell(i, 1, a).font = LABEL_FONT
        ws.cell(i, 1).border = THIN
        ws.cell(i, 2, b).alignment = WRAP
        ws.cell(i, 2).border = THIN
        ws.cell(i, 2).font = BODY_FONT
        ws.row_dimensions[i].height = 52
    set_widths(ws, [28, 110, 18, 18])


def add_compare_sheet(wb, headers: list[str], rows: list[list], name: str = "Сравнительная таблица") -> None:
    ws = wb.create_sheet(name)
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 28
    for r, row in enumerate(rows, start=2):
        write_row(ws, r, row, height=92)
    ws.freeze_panes = "B2"
    set_widths(ws, [22, 36, 48, 16, 42, 52, 48, 42])


def add_recs_sheet(wb, add_items: list[str], dont_items: list[str], horizon: list[str]) -> None:
    ws = wb.create_sheet("Рекомендации")
    ws.cell(1, 1, "Что добавить")
    ws.cell(1, 2, "Что не переносить")
    ws.cell(1, 3, "Горизонт для НИР")
    style_header_row(ws, 1, 3)
    n = max(len(add_items), len(dont_items), len(horizon))
    for i in range(n):
        write_row(
            ws,
            i + 2,
            [
                add_items[i] if i < len(add_items) else "",
                dont_items[i] if i < len(dont_items) else "",
                horizon[i] if i < len(horizon) else "",
            ],
            height=56,
        )
    set_widths(ws, [70, 70, 70])


def add_sources_sheet(wb, items: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet("Источники")
    ws.cell(1, 1, "Источник")
    ws.cell(1, 2, "Ссылка")
    style_header_row(ws, 1, 2)
    for i, (name, url) in enumerate(items, start=2):
        ws.cell(i, 1, name).border = THIN
        ws.cell(i, 1).font = BODY_FONT
        cell = ws.cell(i, 2, url)
        cell.border = THIN
        cell.font = LINK_FONT
        if url.startswith("http"):
            cell.hyperlink = url
        ws.row_dimensions[i].height = 22
    set_widths(ws, [70, 90])


CASE_HEADERS = [
    "Страна",
    "Суд / орган",
    "Цитата",
    "Тип",
    "Сюжет",
    "Правовая позиция",
    "Значение для РФ / НИР",
    "Ссылка",
]


def collect_sources(kind: str) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    seen: set[str] = set()
    for c in COUNTRIES:
        case = c["case_edo"] if kind == "edo" else c["case_ot"]
        for name, url in case["sources"]:
            if url not in seen:
                seen.add(url)
                out.append((f"{c['name']}: {name}", url))
    out.append(("ТК РФ, ст. 22.1 (КЭДО и исключения)", "https://www.consultant.ru/document/cons_doc_LAW_34683/"))
    out.append(("ФЗ № 63‑ФЗ об электронной подписи", "https://www.consultant.ru/document/cons_doc_LAW_112701/"))
    return out


def build_edo_xlsx(path: Path) -> Path:
    wb = Workbook()
    add_summary_sheet(
        wb,
        "Анализ судебной практики иностранных государств по ЭДО (в целом)",
        "Семь стран (август 2026): допустимость электронного документа и подписи в суде. "
        "Сравнение с РФ (63‑ФЗ, Пленум ВС № 25, АПК/ГПК). Номера дел не вымышлены.",
        [
            ("Предмет", "Судебная практика об электронных документах, подписи и допустимости e‑доказательств — не охрана труда."),
            ("Страны", "Германия, Беларусь, Сербия, Испания, Ирландия, Франция, Кения."),
            ("Типы карточек", "Прямая практика — опубликованное дело / линия судов. Аналогия в этом файле почти не нужна: все семь карточек ЭДО опираются на реальные решения или устойчивую процессуальную норму."),
            ("Оси сравнения", "1) презумпция vs свободная оценка; 2) канал подачи / возможность проверить подпись; 3) сертификат допустимости (Кения 106B)."),
            (
                "Ключевой вывод",
                "QES/ЭЦП усиливает позицию, но не всегда достаточна: Германия требует проверяемый Zugang, Франция — именно QES для презумпции, Кения — сертификат 106B, Сербия — правильный канал + QES. "
                "В РФ нет LEC 326.4 и нет 106B; ближе линия «достоверность можно установить».",
            ),
        ],
    )
    rows = [edo_row(c) for c in COUNTRIES]
    rows.append(RF_EDO)
    add_compare_sheet(wb, CASE_HEADERS, rows)
    add_recs_sheet(
        wb,
        [
            "Презумпцию для документов с УКЭП / квалифицированной меткой времени и перенос расходов на необоснованно оспаривающего (урок Испании, LEC 326.4).",
            "Правило доставки: документ с УКЭП должен попасть к адресату в проверяемом виде, а не как скан без валидации (BGH VIII ZR 155/23).",
            "Явно зафиксировать, что презумпция надёжности не действует для простой ЭП, пока не доказан квалифицированный уровень (Cass. 24‑21.034).",
            "Развести «электронный документ с ЭП» и «файл, посланный на почту» (Беларусь e‑court / ст. 84 ХПК; Сербия ст. 157а).",
        ],
        [
            "Не копировать кенийский 106B как универсальный ярлык на каждый электронный файл — это убьёт массовый ЭДО.",
            "Не переносить сербский формализм «e‑mail без QES не подан» на все внутренние документы работодателя.",
            "Не считать испанскую мягкость к pantallazos WhatsApp стандартом для юридически значимого ЭДО.",
            "Не выдавать сертификат оператора/провайдера за презумпцию УКЭП без проверки qualified‑статуса.",
        ],
        [
            "Глава 1 / 3: модели доказываемости ЭП, не страны подряд.",
            "Приложение: карточки дел + эта таблица.",
            "Связка с 63‑ФЗ: УКЭП ≈ QES, но без автоматической европейской презумпции.",
            "Связка с 315‑ФЗ (2026): иностранные/международные УКЭП — точечно, не как eIDAS.",
        ],
    )
    add_sources_sheet(wb, collect_sources("edo"))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def build_ot_xlsx(path: Path) -> Path:
    wb = Workbook()
    add_summary_sheet(
        wb,
        "Анализ судебной практики иностранных государств по ЭДО в охране труда",
        "Семь стран (август 2026): доказывание обучения/инструктажа и электронных записей ОТ. "
        "Прямых дел «электронный журнал инструктажа» почти нет — такие карточки помечены как аналогия.",
        [
            ("Предмет", "Судебная и ближайшая практика о доказательстве обучения, инструктажа и мер безопасности, в т.ч. электронных записей ОТ."),
            ("Страны", "Германия, Беларусь, Сербия, Испания, Ирландия, Франция, Кения."),
            ("Аналогия", "Отдельного прецедента об e‑журнале инструктажа не найдено в DE/BY/RS/ES/IE/FR/KE. Карточка берёт ближайшее реальное дело или связку нормы ОТ + процессуальных правил ЭДО."),
            ("Оси сравнения", "1) можно ли вести запись электронно; 2) какая подпись даёт презумпцию; 3) доказывает ли запись обучение по конкретной опасности."),
            (
                "Ключевой вывод",
                "РФ отвечает на вопрос (1) «нет» для журналов и актов НС в КЭДО (ст. 22.1(3) ТК) — это статут, не вывод суда. "
                "Беларусь отвечает «да» нормой п. 35 Инстр. 175. Сербия — «нет» только для Form 6. "
                "ЕС‑страны спорят про содержание обучения (BAG, Redmond, Cass. 22‑15.624), а не про носитель журнала.",
            ),
        ],
    )
    rows = [ot_row(c) for c in COUNTRIES]
    rows.append(RF_OT)
    add_compare_sheet(wb, CASE_HEADERS, rows)
    add_recs_sheet(
        wb,
        [
            "Легализовать электронный журнал при идентификации, метке времени и защите от правок (Беларусь, п. 35 Инстр. 175) — без УКЭП у каждого рабочего.",
            "Привязать электронный инструктаж к оценке рисков / конкретной операции (BAG 1 ABR 104/09; Redmond [2024] IEHC 167).",
            "Связать журнал ОТ с бременем доказать меры после НС (Cass. soc. 22‑15.624) — система должна закрывать факт, дату, тему и лицо.",
            "Если оставлять исключения из ЭДО, делать их точечными, как сербский Form 6, а не широкими (журналы + акты НС).",
        ],
        [
            "Не выдавать LMS‑клик или скан подписи за электронную подпись журнала (Cass. soc. 21‑19.841).",
            "Не копировать немецкий отказ от подписи на инструктаже без российских компенсаторов: журнал у нас исторически главное доказательство.",
            "Не вводить 106B‑подобный сертификат компьютера на каждую отметку инструктажа (Кения).",
            "Не ссылаться на зарубежный «судебный запрет e‑журнала» — такого запрета в карточках нет.",
        ],
        [
            "Глава 2: процессы инструктажа/обучения, не страна ради страны.",
            "Глава 3: ст. 22.1(3) ТК vs п. 35 Инстр. 175 РБ vs Form 6 Сербии.",
            "Приложение: 7 карточек ОТ + эта таблица + пометка «аналогия».",
            "Нормативный, не судебный характер российского запрета носителя.",
        ],
    )
    add_sources_sheet(wb, collect_sources("ot"))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def build_compare_xlsx(path: Path) -> Path:
    wb = Workbook()
    add_summary_sheet(
        wb,
        "Сравнение судебных практик семи стран (ЭДО и ЭДО в ОТ)",
        "Excel‑аналог файла «00_Сравнение_судебных_практик.docx»: все 14 карточек, оси сходств/различий, сравнение с РФ.",
        [
            ("Объём", "14 карточек: по каждой из семи стран — ЭДО в целом и ЭДО/доказывание в ОТ."),
            ("Метод", "Прямая практика vs аналогия. Номера дел проверены; аналогия явно помечена."),
            ("Похожи", "Суды отделяют «файл существует» от «файл доказывает нужный факт». Скан/клик без идентификации нигде не равен QES/ЭЦП. Нет судебного правила «e‑форма инструктажа всегда ничтожна»."),
            ("Различаются", "Презумпция (ES LEC 326.4; FR только QES) vs свободная оценка (IE, pantallazos). Канал подачи (RS, KE 106B, DE Zugang). ОТ: норма BY/RS vs содержание обучения DE/IE/FR."),
            ("Vs РФ", "Ст. 22.1(3) ТК — статутный запрет носителя. Среди 14 карточек нет судебного подтверждения, что электронный журнал ОТ всегда недопустим."),
        ],
    )
    all_rows = []
    for c in COUNTRIES:
        e, o = c["case_edo"], c["case_ot"]
        all_rows.append(
            [
                c["name"],
                "ЭДО",
                analog_label(e),
                e["court"],
                e["cite"],
                e["title"],
                e["holding"],
                meaning_text(e),
                case_url(e),
            ]
        )
        all_rows.append(
            [
                c["name"],
                "ОТ",
                analog_label(o),
                o["court"],
                o["cite"],
                o["title"],
                o["holding"],
                meaning_text(o),
                case_url(o),
            ]
        )
    add_compare_sheet(
        wb,
        [
            "Страна",
            "Срез",
            "Тип",
            "Суд / орган",
            "Цитата",
            "Сюжет",
            "Правовая позиция",
            "Значение для НИР",
            "Ссылка",
        ],
        all_rows,
        name="Все_14_практик",
    )
    ws = wb.create_sheet("Похожи_и_различаются")
    ws["A1"] = "Ось"
    ws["B1"] = "В чём похожи"
    ws["C1"] = "В чём различаются"
    style_header_row(ws, 1, 3)
    axes = [
        (
            "A. Презумпция vs оценка",
            "Квалифицированная / advanced подпись везде сильнее простой переписки",
            "Испания (LEC 326.4 + штраф) и Франция (только QES, Cass. 2026) стоят на презумпции по‑разному. Ирландия и WhatsApp‑линия Испании — свободная оценка. Кения формально имеет презумпции Evidence Act, но вход перекрыт 106B",
        ),
        (
            "B. Канал / проверяемость",
            "«Отправили файл» ≠ «документ дошёл в юридической форме»",
            "Сербия: без QES и не на тот адрес — не подан. Кения: без 106B — снят. Германия: QES должна быть проверяема адресатом (Zugang). Беларусь: ЭЦП = письменное доказательство по ст. 84 ХПК",
        ),
        (
            "C. Охрана труда",
            "После травмы/НС работодатель должен уметь доказать обучение и меры",
            "Беларусь закрывает журнал нормой (можно). Сербия — нормой Form 6 (нельзя). Германия и Ирландия — содержанием обучения. Франция — бременем мер L.4121. Кения — OSHA s.99 + риск 106B",
        ),
        (
            "D. Носитель журнала",
            "Ни один суд не сказал, что электронная форма инструктажа всегда ничтожна",
            "Запреты, где есть, статутные: РФ ст. 22.1(3) ТК; Сербия Form 6. ЕС‑четвёрка запрета не знает",
        ),
    ]
    for i, row in enumerate(axes, start=2):
        write_row(ws, i, list(row), height=80)
    set_widths(ws, [28, 55, 85])

    ws2 = wb.create_sheet("Сравнение_с_РФ")
    ws2["A1"] = "Вопрос"
    ws2["B1"] = "РФ"
    ws2["C1"] = "Семь стран (по карточкам)"
    style_header_row(ws2, 1, 3)
    rf_rows = [
        (
            "Откуда запрет e‑журнала ОТ",
            "Ст. 22.1 ч. 3 ТК — статут",
            "Судебного запрета нет. BY — прямое разрешение нормой. RS — бумажный Form 6. DE/ES/IE/FR — содержание обучения",
        ),
        (
            "Презумпция ЭП",
            "Сильнее у УКЭП; нет штрафа за оспаривание как LEC 326.4",
            "FR 2026: презумпция только у QES. ES: презумпция квалифицированного сервиса. KE: вход через 106B",
        ),
        (
            "Скан / картинка подписи",
            "На практике слабое доказательство; отдельного «антискана» как Cass. 21‑19.841 нет в кодексе",
            "Франция прямо: скан ≠ электронная подпись (1367), но не всегда «нет подписи»",
        ),
        (
            "Допустимость e‑записи в суде",
            "АПК/ГПК + Пленум № 25: если достоверность устанавливается",
            "BY ст. 84 ХПК прямо включает эл. документы. KE без 106B снимает даже «живой» файл",
        ),
        (
            "Что брать в гл. 3",
            "Точка отсчёта — 63‑ФЗ, 22.1 ТК, 2464, 315‑ФЗ",
            "BY п. 35 — положительная модель журнала. FR/ES — презумпция и антискан. DE Zugang — для уведомлений, не обязательно для внутренней отметки",
        ),
    ]
    for i, row in enumerate(rf_rows, start=2):
        write_row(ws2, i, list(row), height=72)
    set_widths(ws2, [32, 50, 85])

    ws3 = wb.create_sheet("Легенда")
    ws3["A1"] = "Легенда"
    ws3["A1"].font = TITLE_FONT
    ws3["A3"] = "Понятие"
    ws3["B3"] = "Смысл"
    style_header_row(ws3, 3, 2)
    notes = [
        ("Прямая практика", "Опубликованное дело или устойчивая линия судов об электронном документе / подписи / допустимости."),
        ("Аналогия", "Нет дела об электронном журнале ОТ; взята ближайшая практика об обучении, бремени мер или форме подписи."),
        ("Не вымышлено", "BGH VIII ZR 155/23; BAG 1 ABR 104/09; Cass. 24‑21.034; Cass. soc. 22‑15.624 и 21‑19.841; [2025] KEELC 392; [2024] IEHC 167; [2010] IEHC 47; Gž 1664/20 и др."),
        ("РФ в таблицах", "Ориентир для сравнения, не «восьмая зарубежная карточка»."),
    ]
    for i, row in enumerate(notes, start=4):
        write_row(ws3, i, list(row), height=48)
    set_widths(ws3, [24, 110])

    srcs = collect_sources("edo") + collect_sources("ot")
    dedup: list[tuple[str, str]] = []
    seen: set[str] = set()
    for name, url in srcs:
        if url not in seen:
            seen.add(url)
            dedup.append((name, url))
    add_sources_sheet(wb, dedup)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def build_all(extra_dirs: list[Path] | None = None) -> list[Path]:
    paths = [
        build_edo_xlsx(ROOT_EDO),
        build_ot_xlsx(ROOT_OT),
        build_compare_xlsx(ROOT_CMP),
    ]
    import shutil

    for d in extra_dirs or []:
        d.mkdir(parents=True, exist_ok=True)
        for p in paths:
            shutil.copy2(p, d / p.name)
    return paths


if __name__ == "__main__":
    out = build_all()
    print("wrote", [p.name for p in out])
