#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import re
import unicodedata
from openpyxl.styles import PatternFill, Font, Alignment
from difflib import SequenceMatcher
from collections import defaultdict

bank_path = (
    "/home/ubuntu/.cursor/projects/workspace/uploads/"
    "______________________________________2026_31.03__8fc5.xlsx"
)
prev = (
    "/home/ubuntu/.cursor/projects/workspace/uploads/"
    "______________________________________2026_31.03__0b99.xlsx"
)


def normalize(text):
    if text is None:
        return ""
    s = unicodedata.normalize("NFKC", str(text))
    s = s.replace("\xa0", " ").replace("\u200b", "")
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s).rstrip("?.!;: ")
    return s


def extract_questions(ws, sheet_name):
    qs = []
    r = 1
    while r <= ws.max_row:
        v = ws.cell(r, 1).value
        if isinstance(v, str) and re.match(r"^Вопрос\s+\d+", v.strip()):
            m = re.match(r"^Вопрос\s+(\d+)", v.strip())
            num = int(m.group(1)) if m else None
            qrow = r + 1
            while qrow <= ws.max_row and not ws.cell(qrow, 1).value:
                qrow += 1
            q = ws.cell(qrow, 1).value
            if q and isinstance(q, str) and not re.match(r"^Вопрос\s+\d+", q.strip()):
                qs.append(
                    {
                        "sheet": sheet_name,
                        "num": num,
                        "raw": str(q).strip(),
                        "norm": normalize(q),
                    }
                )
            r = qrow + 1
        else:
            r += 1
    return qs


STOP = {
    "и",
    "в",
    "на",
    "по",
    "с",
    "со",
    "к",
    "ко",
    "о",
    "об",
    "от",
    "для",
    "при",
    "из",
    "за",
    "что",
    "как",
    "какой",
    "какая",
    "какое",
    "какие",
    "каким",
    "каком",
    "какую",
    "кого",
    "кому",
    "или",
    "не",
    "ни",
    "а",
    "но",
    "да",
    "ли",
    "же",
    "бы",
    "то",
    "это",
    "эти",
    "том",
    "того",
    "соответствии",
    "согласно",
    "правилам",
    "правило",
    "технической",
    "эксплуатации",
    "тепловых",
    "энергоустановок",
    "объектов",
    "теплоснабжения",
    "теплопотребляющих",
    "установок",
    "должен",
    "должна",
    "должны",
    "быть",
    "может",
    "можно",
    "перечисленного",
    "представленного",
    "перечня",
}


def tokens(n):
    words = re.findall(r"[а-яёa-z0-9\-]+", n, flags=re.I)
    return [w for w in words if len(w) > 2 and w not in STOP]


def score_pair(n1, n2):
    t1, t2 = tokens(n1), tokens(n2)
    seq = SequenceMatcher(None, n1, n2).ratio()
    seq_h = SequenceMatcher(None, n1[:120], n2[:120]).ratio()
    sa, sb = set(t1), set(t2)
    jac = (len(sa & sb) / len(sa | sb)) if sa and sb else 0
    rare = [w for w in (sa & sb) if len(w) >= 6]
    bonus = min(0.15, 0.03 * len(rare))
    return 0.30 * seq + 0.25 * seq_h + 0.45 * jac + bonus, jac, rare


wb = openpyxl.load_workbook(bank_path, data_only=True)
qs_iii = extract_questions(wb["III"], "III")

FA_KEYS = [
    "первой помощи",
    "первую помощь",
    "сердечно-легоч",
    "кровотечен",
    "пострадавш",
    "реанимац",
    "минздрава",
]
qs_fa = []
for name in wb.sheetnames:
    for q in extract_questions(wb[name], name):
        if any(k in q["norm"] for k in FA_KEYS):
            qs_fa.append(q)

fa_by_norm = {}
for q in qs_fa:
    if q["norm"] not in fa_by_norm or q["sheet"] == "III":
        fa_by_norm[q["norm"]] = q
qs_fa_u = list(fa_by_norm.values())

print("III:", len(qs_iii), "FA unique:", len(qs_fa_u))

wb1 = openpyxl.load_workbook(prev)
list1 = []
for r in range(1, wb1["Лист1"].max_row + 1):
    q = wb1["Лист1"].cell(r, 1).value
    if q:
        list1.append(
            {
                "raw": str(q).strip(),
                "norm": normalize(q),
                "code": wb1["Лист1"].cell(r, 2).value,
                "typ": wb1["Лист1"].cell(r, 3).value,
            }
        )


def is_fa(text):
    t = text.lower()
    return any(k in t for k in FA_KEYS)


def find_exact(n, pool):
    for q in pool:
        if q["norm"] == n:
            return q
    for q in pool:
        ln = q["norm"]
        if len(n) >= 40 and (
            n in ln or ln.startswith(n) or ln in n or n.startswith(ln)
        ):
            return q
        if len(n) >= 50 and len(ln) >= 50 and n[:80] == ln[:80]:
            return q
    return None


# Ручные якоря: старый ключ -> предпочтительный № вопроса в III
KEYWORD_PREF = [
    (["температур", "изоляц"], 8),
    (["температур", "поверхност", "изоляц"], 8),
    (["обходов", "осмотров", "рабочих мест"], 10),
    (["стажиров"], 60),
    (["производственн", "инструктаж"], 72),
    (["повторный инструктаж"], 72),
    (["внеочередн", "проверк", "знаний"], 66),
    (["внеочередн", "освидетельствован"], 178),
    (["подготовк", "отопительн"], 225),
    (["режимно-наладоч"], 195),
    (["запорн", "арматур", "горячего"], 236),
    (["пробным давлением"], 250),
    (["испытания на прочность и плотность"], 250),
]


def prefer_by_keywords(old_raw, pool):
    low = old_raw.lower()
    for keys, num in KEYWORD_PREF:
        if all(k in low for k in keys):
            for q in pool:
                if q.get("num") == num:
                    return q
    return None


results = []
for item in list1:
    pool = qs_fa_u if is_fa(item["raw"]) else qs_iii
    ex = find_exact(item["norm"], pool)
    if ex:
        results.append(
            {
                **item,
                "status": "exact",
                "new": ex["raw"],
                "sheet": ex["sheet"],
                "num": ex["num"],
                "score": 1.0,
                "note": "Совпадает с банком 2026",
                "rare": [],
            }
        )
        continue

    pref = prefer_by_keywords(item["raw"], pool) if pool is qs_iii else None

    best = None
    best_sc = 0
    best_jac = 0
    best_rare = []
    for q in pool:
        sc, jac, rare = score_pair(item["norm"], q["norm"])
        if pref and q["num"] == pref["num"]:
            sc += 0.25  # boost keyword-preferred question
        if sc > best_sc:
            best_sc, best_jac, best_rare, best = sc, jac, rare, q

    # If keyword pref exists and is close enough, force as analog
    if pref and best and best["num"] == pref["num"] and best_sc >= 0.30:
        best_sc = max(best_sc, 0.55)

    topic_rare = [
        w
        for w in best_rare
        if w
        not in {
            "энергии",
            "тепловой",
            "теплоносителя",
            "оборудования",
            "наличие",
            "работников",
            "персонала",
            "проводится",
            "случае",
            "необходимо",
            "периодичностью",
            "проводить",
            "требования",
            "предъявляются",
            "исправное",
            "состояние",
            "безопасную",
            "эксплуатацию",
            "ответственный",
            "ответственным",
        }
    ]

    is_analog = (
        best_sc >= 0.55
        or (best_jac >= 0.45 and best_sc >= 0.45)
        or (len(topic_rare) >= 3 and best_sc >= 0.38)
        or (len(topic_rare) >= 2 and best_sc >= 0.48)
    )

    if is_analog:
        st = "analog"
        note = f'Ближайший аналог в листе {best["sheet"]} (сходство {best_sc:.0%})'
        new = best["raw"]
        sheet = best["sheet"]
        num = best["num"]
    elif best_sc >= 0.35:
        st = "none"
        note = (
            "Прямого аналога нет. Показан ближайший по словам из III — только для справки."
        )
        new = best["raw"]
        sheet = best["sheet"]
        num = best["num"]
    else:
        st = "none"
        note = "Актуальной формулировки в банке 2026 не найдено"
        new = ""
        sheet = ""
        num = ""

    results.append(
        {
            **item,
            "status": st,
            "new": new,
            "sheet": sheet,
            "num": num,
            "score": best_sc,
            "note": note,
            "rare": best_rare,
        }
    )

counts = defaultdict(int)
for r in results:
    counts[r["status"]] += 1
print("Counts", dict(counts))

out = openpyxl.Workbook()
leg = out.active
leg.title = "Легенда"
ws = out.create_sheet("Лист1_старые_и_новые", 1)

green = PatternFill("solid", fgColor="C6EFCE")
yellow = PatternFill("solid", fgColor="FFEB9C")
gray = PatternFill("solid", fgColor="D9D9D9")
fills = {"exact": green, "analog": yellow, "none": gray}
labels = {
    "exact": "Есть в банке 2026",
    "analog": "Аналог в банке 2026",
    "none": "Аналог не найден",
}

headers = [
    "№",
    "Вопрос (Лист1, старая формулировка)",
    "Код",
    "Тип",
    "Статус",
    "Новая актуальная формулировка (из файла 31.03.2026)",
    "Лист",
    "№ в банке",
    "Сходство",
    "Комментарий",
]
for c, h in enumerate(headers, 1):
    ws.cell(1, c, h).font = Font(bold=True)
    ws.cell(1, c).alignment = Alignment(wrap_text=True, vertical="top")

for i, r in enumerate(results, 1):
    fill = fills[r["status"]]
    vals = [
        i,
        r["raw"],
        r["code"],
        r["typ"],
        labels[r["status"]],
        r["new"],
        r["sheet"],
        r["num"],
        round(r["score"], 2) if r["score"] else "",
        r["note"],
    ]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(i + 1, c, v)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if c in (2, 5, 6):
            cell.fill = fill
    ws.row_dimensions[i + 1].height = 55

ws.column_dimensions["B"].width = 50
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 55
ws.column_dimensions["J"].width = 45
ws.row_dimensions[1].height = 30

leg["A1"] = (
    "Колонка B — ваш вопрос из Лист1; колонка F — актуальная формулировка из банка 2026"
)
leg["A2"] = (
    "Источник: Перечень_тестовых_вопросов_с_ответами_2026(31.03).xlsx "
    "(в основном лист III — тепло)"
)
leg["A4"] = "Зелёный"
leg["B4"] = "Формулировка совпадает с банком"
leg["A4"].fill = green
leg["A5"] = "Жёлтый"
leg["B5"] = "Подобран смысловой/словесный аналог из III"
leg["A5"].fill = yellow
leg["A6"] = "Серый"
leg["B6"] = (
    "В банке 2026 прямого аналога нет (старый ПТЭТЭ-вопрос убран или сильно переписан)"
)
leg["A6"].fill = gray
leg["A8"] = (
    f'Всего {len(results)}: точных {counts["exact"]}, '
    f'аналогов {counts["analog"]}, без аналога {counts["none"]}'
)

path = "/workspace/Лист1_со_старыми_и_новыми_формулировками.xlsx"
out.save(path)

print("\nANALOGS:")
for r in results:
    if r["status"] == "analog":
        print("OLD:", r["raw"][:100])
        print(
            "NEW:",
            r["new"][:100],
            f'#{r["num"]} {r["score"]:.2f} rare={r["rare"]}',
        )
        print("---")

print("\nNONE with suggested nearest:")
for r in results:
    if r["status"] == "none" and r["new"]:
        print("OLD:", r["raw"][:90])
        print("NEAR:", r["new"][:90], f'{r["score"]:.2f}')
        print("---")

print("Saved", path)
