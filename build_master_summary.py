#!/usr/bin/env python3
"""Одна сводная таблица НИР: все юрисдикции, ОТ, суды, глоссарий, правки, сроки."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from build_edo_countries import COUNTRIES as FIRST, RECS, RF_BASE, RF_SOURCES
from build_seven_countries_pack import COUNTRIES as SEVEN, RF_OT, RF_OT_SRC
from build_nir_extras import U, ot_matrix_rows

OUT = Path("СВОДНАЯ_ТАБЛИЦА_НИР_ЭДО.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
LINK = Font(color="0563C1", name="Calibri", size=9, underline="single")
BODY = Font(name="Calibri", size=9)
TITLE = Font(bold=True, size=16, color="1F4E79", name="Calibri")
NOTE = Font(name="Calibri", size=10, italic=True, color="535353")

FILLS = {
    "1-й комплект": PatternFill("solid", fgColor="DEEAF6"),
    "2-й комплект": PatternFill("solid", fgColor="E2EFDA"),
    "РФ": PatternFill("solid", fgColor="FFF2CC"),
    "Международное": PatternFill("solid", fgColor="FCE4D6"),
    "Документы ОТ": PatternFill("solid", fgColor="E4DFEC"),
    "Глоссарий": PatternFill("solid", fgColor="DDEBF7"),
    "Практика РФ": PatternFill("solid", fgColor="F8CBAD"),
    "Правка закона": PatternFill("solid", fgColor="F4B183"),
    "Сроки и ПДн": PatternFill("solid", fgColor="C6EFCE"),
    "Оси сравнения": PatternFill("solid", fgColor="D0CECE"),
    "Рекомендации": PatternFill("solid", fgColor="C5D9F1"),
}

HEADERS = [
    "№",
    "Блок",
    "Юрисдикция / предмет",
    "Тема",
    "Показатель",
    "Содержание",
    "Сравнение с РФ / значение для НИР",
    "Тип",
    "Источник",
    "Проверить",
]

FIRST_OT = {
    "Европейский союз": {
        "acts": "Директива 89/391 (документирование) + eIDAS для силы ЭП; нац. детали государств‑членов",
        "legal_force": "Нет общесоюзного запрета e‑журнала. Формат документов гибкий; сила ЭП — eIDAS",
        "signature": "SES/AES/QES по риску документа; QES не обязательна для обычного обучения",
        "feature": "Цифровые РА (OiRA) и e‑обучение допустимы; инспектор смотрит идентификацию, целостность, доступность записи",
        "e_journal": "частично",
        "e_training": "да",
        "overview": [
            "Специального «закона об ЭДО в ОТ» нет: документирование обучения и компетентности выводится из рамки 89/391 и национального права, а сила электронной записи — из eIDAS.",
        ],
        "compare": [
            "В ЕС нет аналога ч. 3 ст. 22.1 ТК: носитель журнала не запрещён союзным правом.",
        ],
        "add": ["Легализовать e‑журналы; типовые шаблоны ОПР (OiRA‑логика)."],
        "dont": ["Не делать QES обязательной для каждой отметки инструктажа."],
        "sources": [
            ("OSH Framework Directive 89/391", "https://osha.europa.eu/en/legislation/directives/the-osh-framework-directive/1"),
            ("eIDAS 910/2014", "https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32014R0910"),
        ],
    },
    "Великобритания": {
        "acts": "HSWA 1974 / MHSWR 1999 (suitable and sufficient RA) + ECA/UK eIDAS; goal‑setting HSE",
        "legal_force": "Электронные РА, LMS и e‑наряды‑допуски — обычная практика, если запись sufficient / reproducible",
        "signature": "ECA s.7: ЭП допустима как доказательство; нет бумажного журнала федерального образца",
        "feature": "Критерий sufficient record: кто / когда / программа / результат",
        "e_journal": "частично",
        "e_training": "да",
        "overview": [
            "HSE придерживается goal‑setting: важны оценка рисков и компетентность, а не единый бумажный журнал.",
        ],
        "compare": ["UK смещает акцент с формализма журнала на достаточность записи — в РФ пока наоборот."],
        "add": ["Критерий sufficient record; e‑наряды‑допуски с блокировкой допуска."],
        "dont": ["Goal‑setting без минимальных реквизитов записи."],
        "sources": [
            ("MHSWR 1999, reg. 3", "https://www.legislation.gov.uk/uksi/1999/3242/regulation/3"),
            ("ECA 2000 s.7", "https://www.legislation.gov.uk/ukpga/2000/7/section/7"),
        ],
    },
    "США": {
        "acts": "OSHA LOI 14.08.1997 (electronic training records); 29 CFR Part 1904 + Injury Tracking Application",
        "legal_force": "Electronic training records прямо допустимы. Отчётность по травмам — digital-first (ITA)",
        "signature": "Certification record работодателя; технологическая нейтральность ESIGN",
        "feature": "Прямая спецнорма/разъяснение OSHA по e‑записям обучения",
        "e_journal": "да",
        "e_training": "да",
        "overview": [
            "OSHA 14.08.1997: электронные записи обучения допустимы, если их можно предъявить на инспекции и они подтверждают факт обучения.",
        ],
        "compare": ["США прямо разрешили e‑training records; РФ статутно исключила журналы инструктажа из КЭДО."],
        "add": ["Certification record + audit trail; машиночитаемая отчётность по травмам."],
        "dont": ["Отказ от подтверждения работника только потому, что OSHA делает акцент на certification работодателя."],
        "sources": [
            ("OSHA LOI 14.08.1997", "https://www.osha.gov/laws-regs/standardinterpretations/1997-08-14"),
            ("29 CFR Part 1904", "https://www.osha.gov/laws-regs/regulations/standardnumber/1904"),
        ],
    },
    "Сингапур": {
        "acts": "WSH + Training Records System (TRS); ETA 2010 как общая рамка ЭП",
        "legal_force": "Госреестр обучения: работодатель проверяет статус до допуска",
        "signature": "Подтверждённый training record в TRS, не бумажный журнал инструктажа",
        "feature": "Единственный в выборке государственный сервис проверки обучения подрядчика «здесь и сейчас»",
        "e_journal": "частично",
        "e_training": "да",
        "overview": [
            "MOM Check WSH training records (TRS) позволяет проверить, прошёл ли работник обязательное обучение, до допуска к работе.",
        ],
        "compare": ["В РФ есть реестры обученных по № 2464, но нет удобного сервиса проверки подрядчика как TRS."],
        "add": ["Сервис проверки обучения подрядчиков до допуска."],
        "dont": ["Обязательный тяжёлый госреестр без упрощения для микробизнеса."],
        "sources": [
            ("MOM TRS", "https://www.mom.gov.sg/eservices/services/check-worker-training-records"),
            ("ETA 2010", "https://sso.agc.gov.sg/Act/ETA2010"),
        ],
    },
    "Китай (КНР)": {
        "acts": "Work Safety Law + Закон об электронной подписи («надёжная» ЭП ≈ собственноручная/печать)",
        "legal_force": "Общий закон об ЭП применяется к safety docs; отдельного «закона об ЭДО в ОТ» нет",
        "signature": "«Надёжная» ЭП / лицензированный провайдер",
        "feature": "Единая логика ЭП и для документов ОТ, без отдельного запрета журнала",
        "e_journal": "частично",
        "e_training": "частично",
        "overview": [
            "Крупные предприятия используют electronic safety training и electronic work permits на базе общего режима ЭП.",
        ],
        "compare": ["Китай не выводит журналы инструктажа из электронной формы отдельной кадровой нормой, как ст. 22.1(3) ТК."],
        "add": ["Единая логика ЭП и для инструктажей, без отдельного запрета."],
        "dont": ["Смешивать режим электронных коносаментов с охраной труда."],
        "sources": [
            ("ESL КНР (WIPO Lex)", "https://www.wipo.int/wipolex/en/legislation/details/6563"),
        ],
    },
    "Казахстан / ЕАЭС": {
        "acts": "Закон РК № 370; правила обучения/инструктирования ОТ — электронный журнал с ЭЦП",
        "legal_force": "Прямая спецнорма: электронный журнал инструктажа с ЭЦП инструктирующего и инструктируемого",
        "signature": "ЭЦП обеих сторон записи",
        "feature": "Самый жёсткий положительный образец журнала (ЭЦП); рядом Беларусь — мягче (без обязательной ЭЦП у рабочего)",
        "e_journal": "да",
        "e_training": "да",
        "overview": [
            "Казахстан прямо легализовал электронный журнал инструктажа. Для РФ это образец смены носителя при сохранении содержания инструктажа.",
        ],
        "compare": ["Прямая противоположность ч. 3 ст. 22.1 ТК. Для массового инструктажа белорусский п. 35 практичнее (без ЭЦП у каждого рабочего)."],
        "add": ["Прямой образец легализации e‑журнала для правки ч. 3 ст. 22.1 / № 2464."],
        "dont": ["Требовать УКЭП у каждого рабочего только потому, что в РК стоит ЭЦП."],
        "sources": [
            ("Закон РК № 370", "https://adilet.zan.kz/rus/docs/Z030000370_"),
            ("Правила обучения ОТ РК", "https://zakon.mybuh.kz/rus/docs/v1500012665/"),
        ],
    },
}

GLOSSARY = [
    ("ПЭП", "Простая электронная подпись (63‑ФЗ): коды, пароли, клик", "SES (eIDAS)", U["fz63"][1]),
    ("УНЭП", "Усиленная неквалифицированная ЭП: ключи без квалифицированного сертификата", "AES (eIDAS)", U["fz63"][1]),
    ("УКЭП", "Усиленная квалифицированная ЭП: аккредитованный УЦ, сила собственноручной", "QES (eIDAS)", U["fz63"][1]),
    ("SES / AES / QES", "Простая / усиленная / квалифицированная подпись по eIDAS", "ПЭП / УНЭП / УКЭП", U["eidas"][1]),
    ("ЭЦП (РБ, РК)", "Государственная PKI‑подпись, обычно ближе к УКЭП, без трёх уровней 63‑ФЗ", "УКЭП", U["by175"][1]),
    ("Textform (§ 126b BGB)", "Текстовая форма: указание лица + долговечный носитель, без QES", "сниженный формальный порог, не вид ЭП", "https://www.gesetze-im-internet.de/bgb/__126b.html"),
    ("Скан подписи", "Изображение собственноручной подписи на PDF. Во Франции ≠ электронная подпись (Cass. 21‑19.841)", "не ПЭП/УНЭП/УКЭП", U["eidas"][1]),
    ("УЦ", "Удостоверяющий центр, выдаёт сертификаты ключей ЭП", "аккредитованные УЦ РФ", U["fz63"][1]),
    ("QTSP", "Qualified trust service provider — квалифицированный поставщик услуг доверия в ЕС", "аккредитованный УЦ / доверенная услуга", U["eidas"][1]),
    ("ДТС", "Доверенная третья сторона ЕАЭС: проверяет иностранную ЭЦП между PKI государств‑членов", "Решение ЕЭК № 120", RF_SOURCES[4][1]),
    ("Метка времени", "Доказательство, что файл существовал в данный момент в данном виде", "квалифицированная метка времени eIDAS 2.0", U["eidas"][1]),
    ("Zugang (BGH)", "Доступ волеизъявления: QES должна дойти так, чтобы адресат мог проверить подпись", "доставка УКЭП в проверяемом виде", "https://www.bundesgerichtshof.de/SharedDocs/Entscheidungen/DE/Zivilsenate/VIII_ZS/2023/VIII_ZR_155-23.pdf?__blob=publicationFile&v=1"),
    ("LEC 326.4", "Испанская процессуальная презумпция подлинности для квалифицированной услуги доверия", "в РФ такой презумпции нет", U["eidas"][1]),
    ("Evidence Act s. 106B", "Кенийский сертификат компьютера: без него e‑запись могут снять даже при «живой» подписи", "процессуальный вход, не отраслевой запрет", "https://new.kenyalaw.org/akn/ke/act/1963/46"),
    ("Form 6 (Сербия)", "Учёт обученных безопасной работе и СИЗ; с 2025 г. бумага, остальные формы ОТ — электронно с QES", "точечный запрет, уже чем ст. 22.1(3)", "https://www.paragraf.rs/propisi/zakon-o-bezbednosti-i-zdravlju-na-radu.html"),
    ("Unterweisung", "Немецкий инструктаж по ArbSchG § 12: содержание и связь с оценкой рисков важнее подписи в журнале", "не аналог «журнала как документа»", "https://www.gesetze-im-internet.de/arbschg/__12.html"),
    ("Audit trail", "Журнал событий системы: кто, когда, что изменил", "без него e‑запись слабее прошнурованной бумаги", U["p2464"][1]),
]

AXES = [
    ["Юридическая сила электронной формы", "Недискриминация (UNCITRAL / eIDAS / ESIGN)", "Аналогично при соблюдении 63‑ФЗ", "сходство", "Усилить воспроизводимость записи (мотив ESIGN)", "—", "§ 3.2.1", U["fz63"][1]],
    ["Виды / иерархия подписи", "ЕС: SES/AES/QES; США: нейтральная ЭП", "ПЭП / УНЭП / УКЭП", "сходство", "Презумпции для «secure»‑УНЭП (Сингапур)", "Отказ от иерархии «как в США»", "§ 3.2.2", U["eidas"][1]],
    ["Инфраструктура доверия", "QTSP vs рынок vs УЦ+ДТС", "Аккредитованные УЦ, операторы ЭДО, УЦ ФНС", "сходство", "Квалиф. архив и метка времени (eIDAS 2.0)", "Полный перенос надзора QTSP", "§ 3.4.5", RF_SOURCES[4][1]],
    ["Кадровый ЭДО", "Часто общий режим + согласие", "Отдельный режим ст. 22.1–22.3 ТК", "расхождение", "Сохранить спецрежим, сузить исключения", "Американскую свободу формы без гарантий", "§ 3.1.2", U["tk"][1]],
    ["Инструктажи по ОТ электронно", "Обычно допустимы (OSHA; BY п. 35; KZ; ЕС/UK)", "Ч. 3 ст. 22.1: исключены из КЭДО", "расхождение", "Электронный журнал по модели Беларуси п. 35", "Клик без идентификации", "§ 3.3.1", U["by175"][1]],
    ["Акты / учёт НС", "Электронный учёт ≠ отмена расследования", "Акты НС исключены из КЭДО", "расхождение", "Отдельный контур, не «лёгкий» журнал", "Вывод акта НС в чат без гарантий", "§ 3.3.1", U["oog97"][1]],
    ["Обучение / проверка знания", "E‑records + иногда госреестр (TRS)", "ПП 2464 п. 91–93: протокол уже можно электронно", "частично сходство", "Сервис проверки до допуска (TRS)", "Тяжёлый госреестр для микробизнеса", "§ 3.4.3", U["p2464"][1]],
    ["Оценка профрисков", "Документированная РА + OiRA / BAG", "ОПР есть, единого e‑стандарта нет", "пробел РФ", "Электронные шаблоны ОПР", "Goal‑setting без реквизитов", "§ 3.4.4", "https://www.bundesarbeitsgericht.de/entscheidung/1-abr-104-09/"],
    ["Наряд-допуск", "E‑PTW с блокировками (UK, BY)", "Не в списке 22.1(3); на практике часто бумага", "пробел РФ", "E‑наряды с блокировкой без обучения", "—", "§ 3.4.4", U["by175"][1]],
    ["Ознакомление с ЛНА по ОТ", "Electronic acknowledgement при идентификации", "КЭДО допустим вне исключений 22.1(3)", "сходство", "Единые требования к e‑ознакомлению", "Смешение LMS‑клика и ознакомления", "§ 2.4", U["tk"][1]],
    ["Трансграничное признание", "Автоматизм ЕС; ДТС ЕАЭС", "Ст. 7 63‑ФЗ; ДТС; 315‑ФЗ", "расхождение", "ДТС до практического B2B", "Изоляция PKI", "§ 3.3.4", RF_SOURCES[4][1]],
    ["Доказывание / архив", "Retention, audit trail, timestamp, registered delivery", "Нет сильного квалиф. архива как в eIDAS 2.0; журналы 45 лет", "пробел РФ", "Реквизиты записи ОТ + audit trail + архивный экспорт", "Обязательный EUDI‑кошелёк сразу", "§ 3.4.2 / 3.7", U["rosarkhiv"][1]],
]

RF_PRACTICE = [
    ["ЭДО в целом", "Пленум ВС № 25, п. 65", "разъяснение", "E‑сообщение допустимо, если можно установить отправителя и адресата", "Нет 106B и нет LEC 326.4; достоверность важнее вида ЭП", U["plenum25"][0], U["plenum25"][1]],
    ["Журнал инструктажа", "Письмо Минтруда 15‑2/В‑1677", "разъяснение органа", "Ст. 22.1–22.3 не применяются к документам об инструктажах", "Запрет носителя применяет администрация", U["v1677"][0], U["v1677"][1]],
    ["Журнал + акт НС", "Письмо Минтруда 14‑6/ООГ‑97", "разъяснение органа", "Только бумага; мотив — нельзя менять задним числом", "Технический ответ — audit trail, не обязательно бумага", U["oog97"][0], U["oog97"][1]],
    ["Журнал (ГИТ)", "Онлайнинспекция 178122", "ответ Роструда", "Журналы нельзя вести в КЭДО, нужны личные подписи", "Подтверждает административную практику", U["online"][0], U["online"][1]],
    ["Содержание журнала", "ВС РФ 06.09.2024 № 49‑АД24‑22‑К6", "аналогия (не про носитель)", "Допуск без повторного инструктажа; в журнале нет ЛНА (п. 87 Правил 2464)", "Суд проверяет реквизиты и факт, как BAG/Redmond", U["vs49"][0], U["vs49"][1]],
    ["Акт НС / уголовные дела", "Пленум ВС № 41 (29.11.2018)", "разъяснение", "Материалы расследования НС исследуются как доказательства", "Не смешивать акт НС с журналом инструктажа", U["plenum41"][0], U["plenum41"][1]],
    ["Протокол знания", "ПП 2464 п. 91–93", "статут", "Протокол можно вести электронно при идентификации", "Внутренний разрыв с запретом журнала", U["p2464"][0], U["p2464"][1]],
]

VARIANT_A = (
    "«Положения настоящей статьи и статей 22.2 и 22.3 настоящего Кодекса применяются к документам, подтверждающим прохождение работником инструктажей по охране труда, если такие документы ведутся в электронной форме с использованием программных средств, которые: "
    "1) однозначно идентифицируют работника и лицо, проводившее инструктаж; "
    "2) фиксируют дату и время внесения записи; "
    "3) исключают несанкционированное изменение записи либо обеспечивают неизменяемый журнал событий (audit trail); "
    "4) позволяют воспроизвести запись в течение сроков хранения, установленных законодательством об архивном деле. "
    "Электронная подпись работника не является обязательной, если идентификация обеспечена иным способом, предусмотренным локальным нормативным актом работодателя и не противоречащим настоящему Кодексу. "
    "На акт о несчастном случае на производстве по установленной форме настоящий абзац не распространяется.»"
)

VARIANT_B = (
    "«Регистрация вводного, первичного, повторного, внепланового и целевого инструктажей по охране труда допускается в электронном виде при соблюдении требований части третьей статьи 22.1 Трудового кодекса Российской Федерации в редакции, допускающей такую регистрацию, либо — до изменения Кодекса — параллельно с бумажным журналом как внутренний учёт работодателя, не заменяющий установленную бумажную форму. "
    "Электронный протокол проверки знания требований охраны труда ведётся в порядке пунктов 91–93 настоящих Правил и не подменяет регистрацию инструктажа на рабочем месте.»"
)


def join(items) -> str:
    if not items:
        return ""
    if isinstance(items, str):
        return items
    return " ".join(str(x).strip() for x in items if x)


def src0(sources):
    if not sources:
        return "", ""
    name, url = sources[0][0], sources[0][1]
    return name, url


def add(rows: list, block: str, subject: str, topic: str, metric: str, text: str, vs_rf: str, kind: str, source: str, url: str):
    if not (text or vs_rf):
        return
    rows.append([block, subject, topic, metric, text or "", vs_rf or "", kind, source or "", url or ""])


def flatten_edo(rows, subject: str, pack: str, d: dict, extra_vs: str = ""):
    name, url = src0(d.get("sources"))
    add(rows, pack, subject, "ЭДО в целом", "Основные акты", d.get("acts", ""), extra_vs, "норма", name, url)
    add(rows, pack, subject, "ЭДО в целом", "Юридическая сила", d.get("legal_force", ""), extra_vs, "норма", name, url)
    add(rows, pack, subject, "ЭДО в целом", "Модель подписи", d.get("signature", ""), extra_vs, "норма", name, url)
    add(rows, pack, subject, "ЭДО в целом", "УЦ / доверие", d.get("trust", ""), extra_vs, "норма", name, url)
    add(rows, pack, subject, "ЭДО в целом", "Трансграничность", d.get("crossborder", ""), extra_vs, "норма", name, url)
    add(rows, pack, subject, "ЭДО в целом", "Особенность модели", d.get("feature", ""), extra_vs, "синтез", name, url)
    for p in d.get("overview") or []:
        add(rows, pack, subject, "ЭДО в целом", "Обзор", p, extra_vs, "обзор", name, url)
    for p in d.get("features") or []:
        add(rows, pack, subject, "ЭДО в целом", "Особенность", p, extra_vs, "особенность", name, url)
    for p in d.get("compare") or []:
        add(rows, pack, subject, "ЭДО в целом", "Сравнение с РФ", p, p, "сравнение", name, url)
    for p in d.get("add") or []:
        add(rows, pack, subject, "ЭДО в целом", "Что взять", p, "предложение de lege ferenda / de lege lata", "взять", name, url)
    for p in d.get("dont") or []:
        add(rows, pack, subject, "ЭДО в целом", "Что не переносить", p, "ограничение заимствования", "не брать", name, url)
    for sname, surl in d.get("sources") or []:
        add(rows, pack, subject, "ЭДО в целом", "Источник", sname, "", "источник", sname, surl)


def flatten_ot(rows, subject: str, pack: str, d: dict):
    name, url = src0(d.get("sources"))
    add(rows, pack, subject, "ЭДО в охране труда", "Акты ОТ", d.get("acts", ""), "", "норма", name, url)
    add(rows, pack, subject, "ЭДО в охране труда", "Можно ли e‑журнал / e‑учёт", d.get("legal_force", ""), "", "норма", name, url)
    add(rows, pack, subject, "ЭДО в охране труда", "Подпись / форма", d.get("signature", ""), "", "норма", name, url)
    if d.get("e_journal"):
        add(rows, pack, subject, "ЭДО в охране труда", "E‑журнал (кратко)", d["e_journal"], "", "статус", name, url)
    if d.get("e_training"):
        add(rows, pack, subject, "ЭДО в охране труда", "E‑обучение (кратко)", d["e_training"], "", "статус", name, url)
    add(rows, pack, subject, "ЭДО в охране труда", "Особенность ОТ", d.get("feature", ""), "", "синтез", name, url)
    add(rows, pack, subject, "ЭДО в охране труда", "Доверие / система", d.get("trust", ""), "", "норма", name, url)
    add(rows, pack, subject, "ЭДО в охране труда", "Трансграничность ОТ", d.get("crossborder", ""), "", "норма", name, url)
    for p in d.get("overview") or []:
        add(rows, pack, subject, "ЭДО в охране труда", "Обзор", p, "", "обзор", name, url)
    for p in d.get("features") or []:
        add(rows, pack, subject, "ЭДО в охране труда", "Особенность", p, "", "особенность", name, url)
    for p in d.get("compare") or []:
        add(rows, pack, subject, "ЭДО в охране труда", "Сравнение с РФ", p, p, "сравнение", name, url)
    for p in d.get("add") or []:
        add(rows, pack, subject, "ЭДО в охране труда", "Что взять", p, "предложение для гл. 3", "взять", name, url)
    for p in d.get("dont") or []:
        add(rows, pack, subject, "ЭДО в охране труда", "Что не переносить", p, "ограничение заимствования", "не брать", name, url)
    for sname, surl in d.get("sources") or []:
        add(rows, pack, subject, "ЭДО в охране труда", "Источник", sname, "", "источник", sname, surl)


def flatten_case(rows, subject: str, pack: str, slice_name: str, case: dict):
    name, url = src0(case.get("sources"))
    analog = "аналогия" if case.get("analog") else "прямая практика"
    topic = f"Суд: {slice_name}"
    add(rows, pack, subject, topic, "Тип карточки", analog, "Если аналогия — нет дела именно об e‑журнале ОТ", analog, name, url)
    add(rows, pack, subject, topic, "Суд", case.get("court", ""), "", analog, name, url)
    add(rows, pack, subject, topic, "Цитата", case.get("cite", ""), "", analog, name, url)
    add(rows, pack, subject, topic, "Сюжет", case.get("title", ""), "", analog, name, url)
    add(rows, pack, subject, topic, "Фабула", case.get("facts", ""), "", analog, name, url)
    add(rows, pack, subject, topic, "Позиция суда", case.get("holding", ""), "", analog, name, url)
    for p in case.get("meaning") or []:
        add(rows, pack, subject, topic, "Значение для НИР", p, p, analog, name, url)
    for sname, surl in case.get("sources") or []:
        add(rows, pack, subject, topic, "Источник", sname, "", "источник", sname, surl)


def collect_rows() -> list[list]:
    rows: list[list] = []

    add(
        rows,
        "Международное",
        "UNCITRAL",
        "ЭДО в целом",
        "Рамка",
        "MLEC 1996 — недискриминация электронной формы; MLES 2001 — надёжность ЭП; Конвенция 2005 — международные электронные договоры; MLETR 2017 — электронные коносаменты. К 2026 MLETR: 13 юрисдикций, в т.ч. SG, UK, FR, CN.",
        "РФ опирается на ту же логику недискриминации в 63‑ФЗ, но MLETR в охрану труда не переносится.",
        "норма",
        "UNCITRAL MLEC",
        "https://uncitral.un.org/en/texts/ecommerce/modellaw/electronic_commerce",
    )
    for name, url in [
        ("MLES 2001", "https://uncitral.un.org/en/texts/ecommerce/modellaw/electronic_signatures"),
        ("Конвенция 2005", "https://uncitral.un.org/en/texts/ecommerce/conventions/electronic_communications"),
        ("MLETR + статус", "https://uncitral.un.org/en/texts/ecommerce/modellaw/electronic_transferable_records/status"),
    ]:
        add(rows, "Международное", "UNCITRAL", "ЭДО в целом", "Источник", name, "", "источник", name, url)

    for c in FIRST:
        flatten_edo(rows, c["short"], "1-й комплект", c)
        ot = FIRST_OT.get(c["short"])
        if ot:
            flatten_ot(rows, c["short"], "1-й комплект", ot)

    for c in SEVEN:
        flatten_edo(rows, c["name"], "2-й комплект", c["edo"])
        flatten_ot(rows, c["name"], "2-й комплект", c["ot"])
        flatten_case(rows, c["name"], "2-й комплект", "ЭДО", c["case_edo"])
        flatten_case(rows, c["name"], "2-й комплект", "ОТ", c["case_ot"])

    rf_edo = {
        "acts": "63‑ФЗ; ст. 22.1–22.3 ТК; 315‑ФЗ (04.08.2026); Решение ЕЭК № 120",
        "legal_force": "Электронный документ с ЭП признаётся при соблюдении вида подписи и закона/соглашения",
        "signature": "ПЭП / УНЭП / УКЭП (иерархия близка eIDAS SES/AES/QES)",
        "trust": "Аккредитованные УЦ, операторы ЭДО, УЦ ФНС",
        "crossborder": "Ст. 7 63‑ФЗ, договоры сторон, ДТС ЕАЭС; 315‑ФЗ — УКЭП международных организаций в РФ",
        "feature": "Отдельный КЭДО; ч. 3 ст. 22.1 выводит журналы инструктажа и акты НС",
        "overview": RF_BASE,
        "features": [],
        "compare": [],
        "add": RECS["add"],
        "dont": RECS["dont"],
        "sources": RF_SOURCES,
    }
    flatten_edo(rows, "Российская Федерация", "РФ", rf_edo)
    rf_ot = {
        "acts": "ТК ст. 22.1 ч. 3; ПП 2464; письма Минтруда 15‑2/В‑1677 и 14‑6/ООГ‑97",
        "legal_force": "Журналы инструктажа и акты НС исключены из КЭДО. Протокол проверки знания (п. 91–93 Правил № 2464) уже можно вести электронно",
        "signature": "На практике бумажный журнал + собственноручная подпись. УКЭП у рабочего не требуется законом — требуется бумага",
        "trust": "Бумажный журнал + ПП 2464; операторы КЭДО — для иных кадровых документов",
        "crossborder": "Не применимо к журналу инструктажа, пока действует запрет носителя",
        "feature": "Уникальный в выборке статутный запрет носителя. Суд проверяет содержание журнала, а не «электронность»",
        "e_journal": "нет",
        "e_training": "частично (протокол — да, журнал — нет)",
        "overview": RF_OT,
        "compare": ["Запрет носителя — законодательный выбор, не вывод суда о ничтожности электронной формы."],
        "add": [
            "Вариант A: абзац к ч. 3 ст. 22.1 по модели Беларуси п. 35 (идентификация + время + целостность, без УКЭП у каждого рабочего).",
            "Акт НС не смешивать с журналом.",
        ],
        "dont": ["Не требовать УКЭП у каждого рабочего.", "Не копировать 106B как условие каждой отметки."],
        "sources": RF_OT_SRC + [U["oog97"], U["p2464"]],
    }
    flatten_ot(rows, "Российская Федерация", "РФ", rf_ot)

    for doc, rf, abroad, check, url in ot_matrix_rows():
        add(rows, "Документы ОТ", doc, "Процесс ОТ", "РФ: можно ли электронно", rf, check, "матрица", "см. URL", url)
        add(rows, "Документы ОТ", doc, "Процесс ОТ", "За рубежом", abroad, check, "матрица", "см. URL", url)
        add(rows, "Документы ОТ", doc, "Процесс ОТ", "Что проверяет суд / инспектор", check, check, "матрица", "см. URL", url)

    for term, sense, analog, url in GLOSSARY:
        add(rows, "Глоссарий", term, "Термин", "Смысл", sense, f"Ближайший аналог: {analog}", "термин", term, url)

    for slice_, act, typ, pos, meaning, src, url in RF_PRACTICE:
        add(rows, "Практика РФ", "Российская Федерация", slice_, act, pos, meaning, typ, src, url)

    add(rows, "Правка закона", "ст. 22.1 ТК / ПП 2464", "De lege ferenda", "Зачем две правки", "Журнал — регистр допуска (закрывается идентификацией, временем, audit trail). Акт НС — процессуальный документ (мотив Минтруда: нельзя менять задним числом). Не освобождать одной фразой.", "Акт НС оставить на отдельном режиме", "предложение", U["oog97"][0], U["oog97"][1])
    add(rows, "Правка закона", "ст. 22.1 ТК", "Вариант A (предпочтительный)", "Текст абзаца", VARIANT_A, "Образец условий 1–3 — п. 35 Инструкции РБ № 175. На акт НС абзац не распространяется.", "проект нормы", U["tk"][0], U["tk"][1])
    add(rows, "Правка закона", "ПП 2464", "Вариант B (без правки ТК)", "Текст", VARIANT_B, "Пока 22.1(3) не изменена, B не заменяет бумажный журнал — только фиксирует различие протокола и журнала.", "проект нормы", U["p2464"][0], U["p2464"][1])
    add(rows, "Правка закона", "ст. 22.1 / 2464", "Не включать", "Ограничения", "Не распространять лёгкий e‑журнал на акт НС; не требовать УКЭП у каждого рабочего; не копировать 106B; не отменять подтверждение работника по немецкой модели без компенсаторов.", "границы заимствования", "предложение", U["by175"][0], U["by175"][1])
    add(rows, "Правка закона", "ЛНА / 2464", "Реквизиты e‑записи", "Минимум", "Идентификаторы сторон; дата, время, вид; наименование ЛНА (п. 87 Правил № 2464; ВС 49‑АД24‑22‑К6); привязка к рабочему месту / ОПР; audit trail; хранение 45 лет (Росархив № 236 п. 423 «а»).", "без этих реквизитов цифровизация хуже бумаги", "предложение", U["rosarkhiv"][0], U["rosarkhiv"][1])

    add(rows, "Сроки и ПДн", "Журнал инструктажа", "Хранение", "РФ", "Приказ Росархива № 236 п. 423 «а»: 45 лет", "LMS без архивного экспорта непригоден", "норма", U["rosarkhiv"][0], U["rosarkhiv"][1])
    add(rows, "Сроки и ПДн", "Протокол / журнал проверки знаний", "Хранение", "РФ", "5 лет (п. 423 «б»). Протокол уже может быть электронным (2464 п. 91–93)", "короткий контур, проще журнала", "норма", U["p2464"][0], U["p2464"][1])
    add(rows, "Сроки и ПДн", "Материалы НС / травм", "Хранение", "РФ", "45 лет; при жертвах — постоянно", "не смешивать с e‑журналом инструктажа", "норма", U["rosarkhiv"][0], U["rosarkhiv"][1])
    add(rows, "Сроки и ПДн", "Электронный журнал", "Персональные данные", "152‑ФЗ", "ФИО, должность, рабочее место, время; цель — обязанности по ОТ; минимизация; безопасность; трансграничная передача (ст. 12). Архивные 45 лет задают долгий срок, а не «удалить после увольнения».", "биометрия не равна идентификации по умолчанию; облако иностранного LMS — отдельное решение", "норма", U["pd"][0], U["pd"][1])
    add(rows, "Сроки и ПДн", "Электронный журнал в ЕС", "Персональные данные", "GDPR ст. 5 и 32", "Цель, минимизация, безопасность, сроки хранения. Это не запрет e‑журнала, а требование спроектировать доступ, логи и место сервера.", "сравнительный слой к 152‑ФЗ", "норма", U["gdpr"][0], U["gdpr"][1])
    add(rows, "Сроки и ПДн", "Сербия", "Хранение форм ОТ", "ориентир", "Правилник 5/2025: сроки часто 40 лет; Form 6 на бумаге", "длинный архив как у РФ", "сравнение", "Закон о БЗР РС", "https://www.paragraf.rs/propisi/zakon-o-bezbednosti-i-zdravlju-na-radu.html")

    for axis, abroad, rf, typ, take, dont, para, url in AXES:
        add(rows, "Оси сравнения", axis, "Глава 3", "За рубежом", abroad, f"{typ}. Параграф {para}", typ, para, url)
        add(rows, "Оси сравнения", axis, "Глава 3", "РФ", rf, f"{typ}. Параграф {para}", typ, para, url)
        add(rows, "Оси сравнения", axis, "Глава 3", "Что взять", take, para, "взять", para, url)
        add(rows, "Оси сравнения", axis, "Глава 3", "Что не переносить", dont, para, "не брать", para, url)

    for p in RECS["general"]:
        add(rows, "Рекомендации", "Сводка по ЭДО (1-й комплект)", "Общий вывод", "Синтез", p, "", "синтез", RECS["title"], RF_SOURCES[0][1])
    for p in RECS["matrix"]:
        add(rows, "Рекомендации", "Горизонт внедрения", "ЭДО в целом", "Горизонт", p, "", "горизонт", RECS["title"], RF_SOURCES[0][1])

    return rows


def write_xlsx(rows: list[list]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "СВОДНАЯ"

    ws.merge_cells("A1:J1")
    ws["A1"] = "Сводная таблица НИР: ЭДО, ЭДО в охране труда, судебная практика, правки ТК, сроки и ПДн"
    ws["A1"].font = TITLE
    ws.merge_cells("A2:J2")
    ws["A2"] = (
        "Одна таблица со всем содержимым справок. Включите автофильтр (уже включён). "
        "Фильтр по столбцу «Блок» или «Юрисдикция / предмет». Кликом по столбцу «Проверить» открывается официальный источник. "
        "Номера дел не вымышлены. Тип «аналогия» = нет дела именно об электронном журнале ОТ. Актуализация: август 2026."
    )
    ws["A2"].font = NOTE
    ws["A2"].alignment = WRAP
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 42

    header_row = 4
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(header_row, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN
    ws.row_dimensions[header_row].height = 22

    for i, row in enumerate(rows, start=1):
        r = header_row + i
        values = [i, *row]
        fill = FILLS.get(row[0])
        for c, val in enumerate(values, start=1):
            cell = ws.cell(r, c, val)
            cell.alignment = WRAP
            cell.border = THIN
            cell.font = BODY
            if fill and c <= 3:
                cell.fill = fill
            if c == 10 and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.font = LINK
        ws.row_dimensions[r].height = 48

    last = header_row + len(rows)
    ws.auto_filter.ref = f"A{header_row}:J{last}"
    ws.freeze_panes = "C5"
    widths = [6, 18, 28, 22, 28, 72, 48, 16, 36, 42]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A{header_row}:J{last}"

    ws2 = wb.create_sheet("Как_читать")
    ws2["A1"] = "Как пользоваться сводной таблицей"
    ws2["A1"].font = TITLE
    guide = [
        ("Лист «СВОДНАЯ»", "Это и есть единственная большая таблица. Все справки, суды, глоссарий, правки и сроки сведены в строки."),
        ("Фильтр «Блок»", "1-й комплект / 2-й комплект / РФ / Документы ОТ / Глоссарий / Практика РФ / Правка закона / Сроки и ПДн / Оси сравнения / Рекомендации / Международное."),
        ("Фильтр «Юрисдикция»", "Одна страна целиком: и ЭДО, и охрана труда, и судебные карточки."),
        ("Столбец «Проверить»", "Кликабельный URL официального закона или судебного решения."),
        ("Тип «аналогия»", "Нет опубликованного спора именно об электронном журнале ОТ. Карточка про содержание обучения / форму подписи."),
        ("Главный вывод", "РФ уникальна запретом КЭДО для журналов инструктажа и актов НС (ст. 22.1 ч. 3 ТК). Ближайший образец журнала — Беларусь, п. 35 Инструкции № 175. Протокол знания в РФ уже можно вести электронно (ПП 2464 п. 91–93)."),
        ("Строк в таблице", str(len(rows))),
    ]
    ws2["A3"] = "Вопрос"
    ws2["B3"] = "Ответ"
    for c in (1, 2):
        ws2.cell(3, c).fill = HEADER_FILL
        ws2.cell(3, c).font = HEADER_FONT
        ws2.cell(3, c).border = THIN
    for i, (a, b) in enumerate(guide, start=4):
        ws2.cell(i, 1, a).font = Font(bold=True, name="Calibri", size=10)
        ws2.cell(i, 1).border = THIN
        ws2.cell(i, 2, b).alignment = WRAP
        ws2.cell(i, 2).border = THIN
        ws2.cell(i, 2).font = BODY
        ws2.row_dimensions[i].height = 40
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 110

    wb.save(OUT)
    return OUT


def main():
    rows = collect_rows()
    path = write_xlsx(rows)
    print(f"rows={len(rows)} file={path} size={path.stat().st_size}")


if __name__ == "__main__":
    main()
