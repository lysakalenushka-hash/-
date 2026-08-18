#!/usr/bin/env python3
"""Глоссарий иностранных слов и аббреviatur из сводной таблицы НИР по ЭДО."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from openpyxl.utils.cell import quote_sheetname
from openpyxl.worksheet.hyperlink import Hyperlink as XLHyperlink

from build_edo_countries import (
    RF_SOURCES,
    add_heading,
    add_para,
    add_source_list,
    add_title_block,
    setup_doc,
)
from build_nir_extras import U
from build_seven_countries_pack import add_table, save

OUT_XLSX = Path("Глоссарий_иностранных_терминов_НИР_ЭДО.xlsx")
OUT_DOCX = Path("Глоссарий_иностранных_терминов_НИР_ЭДО.docx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
CAT_FILL = PatternFill("solid", fgColor="D6DCE4")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
LINK = Font(color="0563C1", name="Calibri", size=9, underline="single")
BODY = Font(name="Calibri", size=10)
TITLE = Font(bold=True, size=16, color="1F4E79", name="Calibri")

# term, category, lang, meaning_ru, analog_rf, where_in_table, note, url
ENTRIES: list[tuple[str, str, str, str, str, str, str, str]] = [
    # --- Подпись и ЭДО (общие) ---
    ("eIDAS", "Подпись и ЭДО", "ЕС", "Регламент (EU) № 910/2014 об электронной идентификации и услугах доверия. Задаёт уровни подписи SES/AES/QES и правила их признания внутри ЕС.", "63‑ФЗ об ЭП; иерархия ПЭП / УНЭП / УКЭП", "ЕС, Испания, Франция, Ирландия, Сербия", "Не про журнал инструктажа — только про силу подписи", U["eidas"][1]),
    ("eIDAS 2.0", "Подпись и ЭДО", "ЕС", "Обновление eIDAS — регламент (EU) 2024/1183: европейский кошелёк цифровой идентичности (EUDI Wallet), квалифицированное архивирование, электронные реестры.", "315‑ФЗ, развитие инфраструктуры УЦ", "ЕС, ось «Инфраструктура доверия»", "", U["eidas"][1]),
    ("SES", "Подпись и ЭДО", "ЕС (eIDAS)", "Simple Electronic Signature — простая электронная подпись: логин, SMS, клик «согласен».", "ПЭП по 63‑ФЗ", "Столбец «Модель подписи»", "Не требует сертификата УЦ", U["eidas"][1]),
    ("AES", "Подпись и ЭДО", "ЕС (eIDAS)", "Advanced Electronic Signature — усиленная неквалифицированная подпись: криптоключи, но без квалифицированного сертификата EU QTSP.", "УНЭП", "«Модель подписи»", "", U["eidas"][1]),
    ("QES", "Подпись и ЭДО", "ЕС (eIDAS)", "Qualified Electronic Signature — квалифицированная подпись через аккредитованного поставщика; в ЕС = сила собственноручной.", "УКЭП", "Сербия (Form 6 исключение), Франция, Испания", "Для массового инструктажа обычно не нужна", U["eidas"][1]),
    ("QTSP", "Подпись и ЭДО", "ЕС", "Qualified Trust Service Provider — аккредитованный поставщик квалифицированных услуг доверия (подпись, печать, метка времени, архив).", "Аккредитованный УЦ РФ", "ЕС, eIDAS 2.0", "Trusted lists — публичные реестры QTSP", U["eidas"][1]),
    ("EUDI Wallet", "Подпись и ЭДО", "ЕС", "European Digital Identity Wallet — государственный или одобренный «кошелёк» цифровой личности и атрибутов.", "Пока нет прямого аналога в РФ", "ЕС, eIDAS 2.0", "Не путать с корпоративным SSO", U["eidas"][1]),
    ("PKI", "Подпись и ЭДО", "междун.", "Public Key Infrastructure — инфраструктура открытых ключей: УЦ, сертификаты, цепочки доверия.", "Система аккредитованных УЦ + операторы ЭДО", "ДТС ЕАЭС, трансграничность", "", U["fz63"][1]),
    ("ESIGN Act", "Подпись и ЭДО", "США", "Federal Electronic Signatures in Global and National Commerce Act — запрет отказывать документу только из‑за электронной формы.", "63‑ФЗ, недискриминация электронной формы", "США, ось сравнения", "Не OSHA и не про журнал ОТ", "https://www.law.cornell.edu/uscode/text/15/chapter-96"),
    ("UETA", "Подпись и ЭДО", "США", "Uniform Electronic Transactions Act — модельный закон штатов о равенстве электронных и бумажных записей.", "63‑ФЗ + соглашение сторон", "США", "Дополняет ESIGN на уровне штатов", "https://www.uniformlaws.org/committees/community-home"),
    ("ECA 2000", "Подпись и ЭДО", "Великобритания", "Electronic Communications Act 2000 — section 7: электронная подпись допускается как доказательство; электронность сама по себе не делает её ничтожной.", "63‑ФЗ, ст. 6", "UK: «Подпись / форма»", "Слой ЭДО, не HSE и не MHSWR", "https://www.legislation.gov.uk/ukpga/2000/7/section/7"),
    ("ETA 2010", "Подпись и ЭДО", "Сингапур", "Electronic Transactions Act 2010 — общий режим электронных подписей и записей.", "63‑ФЗ", "Сингапур", "Отдельно от TRS (реестр обучения)", "https://sso.agc.gov.sg/Act/ETA2010"),
    ("ESL / Закон об ЭП КНР", "Подпись и ЭДО", "Китай", "Electronic Signature Law — «надёжная» (reliable) электронная подпись ≈ собственноручная / печать организации.", "УКЭП / усиленная ЭП", "Китай", "", "https://www.wipo.int/wipolex/en/legislation/details/6563"),
    ("Textform", "Подпись и ЭДО", "Германия", "§ 126b BGB — текстовая форма: указание лица + долговечный носитель (e‑mail, PDF), без квалифицированной подписи.", "Не вид ПЭП, а облегчённая письменная форма", "Германия, NachwG 2025", "Снижает порог для трудовых уведомлений", "https://www.gesetze-im-internet.de/bgb/__126b.html"),
    ("Schriftform", "Подпись и ЭДО", "Германия", "Строгая письменная форма по BGB — обычно собственноручная подпись на бумаге или QES.", "Собственноручная подпись / УКЭП где закон требует", "Германия", "Увольнение, срочный договор — остаётся Schriftform", "https://www.gesetze-im-internet.de/bgb/__126.html"),
    ("Zugang", "Подпись и ЭДО", "Германия", "Доступ волеизъявления (BGH): QES должна быть доставлена так, чтобы получатель мог её проверить.", "Доставка УКЭП в проверяемом виде", "Германия, судебная карточка ЭДО", "Дело BGH VIII ZR 155/23", "https://www.bundesgerichtshof.de/SharedDocs/Entscheidungen/DE/Zivilsenate/VIII_ZS/2023/VIII_ZR_155-23.pdf?__blob=publicationFile&v=1"),
    ("click-wrap", "Подпись и ЭДО", "Ирландия", "Согласие с условиями через клик «I agree» на сайте или в приложении.", "ПЭП + доведение условий до пользователя", "Ирландия", "Суд смотрит, были ли условия доведены", "https://www.irishstatutebook.ie/eli/2000/act/27/section/13/enacted/en/html"),
    ("pantallazos", "Подпись и ЭДО", "Испания", "Снимки экрана (WhatsApp и др.); в Испании могут быть доказательством, но слабее QES (STS 116/2025).", "Скрин переписки — не ЭП", "Испания", "Не равно электронной подписи", "https://www.legaltoday.com/practica-juridica/derecho-penal/valor-probatorio-de-los-pantallazos-de-conversaciones-de-whatsapp-o-sus-transcripciones-sts-116-2025-de-13-de-febrero-2025-09-29/"),
    # --- Международные модели ---
    ("UNCITRAL", "Международные модели", "междун.", "Комиссия ООН по праву международной торговли; модели MLEC, MLES, MLETR.", "Логика недискриминации в 63‑ФЗ", "Блок «Международное»", "", "https://uncitral.un.org/"),
    ("MLEC", "Международные модели", "междун.", "Model Law on Electronic Commerce — электронная форма не ухудшает юридическую силу.", "63‑ФЗ", "UNCITRAL", "", "https://uncitral.un.org/en/texts/ecommerce/modellaw/electronic_commerce"),
    ("MLES", "Международные модели", "междун.", "Model Law on Electronic Signatures — надёжность связи подписи с лицом.", "63‑ФЗ", "UNCITRAL", "", "https://uncitral.un.org/en/texts/ecommerce/modellaw/electronic_signatures"),
    ("MLETR", "Международные модели", "междун.", "Model Law on Electronic Transferable Records — электронные коносаменты и иные передаваемые документы.", "Не про охрану труда", "Китай, UK, SG", "Не переносить в журнал инструктажа", "https://uncitral.un.org/en/texts/ecommerce/modellaw/electronic_transferable_records/status"),
    ("Peppol", "Международные модели", "ЕС/B2B", "Сеть обмена электронными счетами и заказами в госзакупках ЕС.", "Операторы ЭДО + ФНС для УПД", "ЕС (B2B)", "Не про ОТ", "https://peppol.org/"),
    # --- Инфраструктура и доказательства ---
    ("audit trail", "Доказательства", "англ.", "Журнал событий системы: кто, когда, что создал или изменил. Без него e‑запись слабее бумаги.", "Журнал событий КЭДО / СУОТ", "Правка закона, UK, США", "Ответ на мотив Минтруда «нельзя менять задним числом»", U["p2464"][1]),
    ("timestamp / метка времени", "Доказательства", "англ./ЕС", "Квалифицированная метка времени — доказательство, что файл существовал в данный момент в данном виде.", "Метка времени (eIDAS 2.0)", "ЕС, ось «Доказывание»", "", U["eidas"][1]),
    ("registered delivery", "Доказательства", "ЕС", "Qualified Registered Electronic Delivery — зарегистрированная доставка юридически значимого сообщения (получено / не получено).", "Зарегистрированное e‑уведомление", "eIDAS 2.0", "", U["eidas"][1]),
    ("qualified archive", "Доказательства", "ЕС", "Квалифицированное электронное архивирование — услуга доверия с доказуемой целостностью на длинный срок.", "Архивное хранение 45 лет (Росархив)", "eIDAS 2.0, сроки", "", U["rosarkhiv"][1]),
    ("retention", "Доказательства", "англ.", "Срок и правила хранения записей (retention policy).", "Сроки по Росархиву № 236", "Сроки и ПДн", "", U["rosarkhiv"][1]),
    ("reproducible", "Доказательства", "англ.", "Запись, которую можно воспроизвести в неизменном виде при проверке.", "Воспроизводимость электронного документа", "UK «Можно ли e‑журнал»", "", "https://www.hse.gov.uk/pubns/indg345.htm"),
    ("Evidence Act s. 106B", "Доказательства", "Кения", "Сертификат компьютера / устройства: процессуальное условие допуска e‑записи в суд Кении.", "Нет прямого аналога; близко — экспертиза носителя", "Кения", "Не отраслевой запрет журнала", "https://new.kenyalaw.org/akn/ke/act/1963/46"),
    ("LEC 326.4", "Доказательства", "Испания", "Презумпция подлинности записи, созданной через квалифицированную услугу доверия.", "В РФ такой презумпции нет", "Испания", "", "https://www.iberley.es/legislacion/articulo-326-ley-enjuiciamiento-civil"),
    ("holding", "Доказательства", "англ. (суд)", "Позиция суда по существу спора (итоговый вывод), не пересказ фактов.", "Мотивировочная часть решения", "Судебные карточки «Позиция суда»", "", U["plenum25"][1]),
    ("certification record", "Доказательства", "США (OSHA)", "Запись работодателя, подтверждающая, что обучение проведено (employer certification).", "Подпись ответственного / запись в журнале", "США OSHA", "Акцент на работодателе, не на QES рабочего", "https://www.osha.gov/laws-regs/standardinterpretations/1997-08-14"),
    # --- Охрана труда UK ---
    ("HSWA 1974", "Охрана труда UK", "Великобритания", "Health and Safety at Work etc. Act 1974 — базовый закон: работодатель обеспечивает безопасность, обучение, информацию.", "ТК РФ, ФЗ об ОТ", "UK «Акты ОТ»", "Не про ЭДО", "https://www.legislation.gov.uk/ukpga/1974/37"),
    ("MHSWR 1999", "Охрана труда UK", "Великобритания", "Management of Health and Safety at Work Regulations 1999 — оценка рисков, обучение, мониторинг.", "ПП 2464, обязанности по ОТ", "UK", "reg. 3: suitable and sufficient — про RA, не журнал", "https://www.legislation.gov.uk/uksi/1999/3242/regulation/3"),
    ("suitable and sufficient", "Охрана труда UK", "англ.", "«Подходящая и достаточная» — стандарт качества оценки рисков (risk assessment) в MHSWR reg. 3.", "Достаточность мер ОТ / ОПР", "UK «Акты ОТ»", "Не synonym «sufficient record»", "https://www.legislation.gov.uk/uksi/1999/3242/regulation/3"),
    ("sufficient record", "Охрана труда UK", "формула НИР", "Рабочий критерий составителя: запись обучения должна отвечать «кто / когда / программа / результат». Не термин HSE и не цитата MHSWR.", "Реквизиты записи инструктажа (предложение для правки ТК)", "UK «Особенность ОТ»", "См. лист «Как_писать_текст» в сводной таблице", "https://www.hse.gov.uk/pubns/indg345.htm"),
    ("goal-setting", "Охрана труда UK", "англ.", "Подход HSE: закон задаёт цель (безопасность, компетентность), а не бланк каждого документа.", "Принцип достаточности мер vs формализм", "UK «Обзор», «Что не переносить»", "Не отменяет минимум реквизитов для РФ", "https://www.hse.gov.uk/pubns/indg345.htm"),
    ("HSE", "Охрана труда UK", "Великобритания", "Health and Safety Executive — инспекция и разъяснения по охране труда.", "ГИТ / Роструд (частично)", "UK", "", "https://www.hse.gov.uk/"),
    ("INDG345", "Охрана труда UK", "Великобритания", "Листовка HSE «Health and safety training» — совет хранить training records и refresher; бланка журнала нет.", "Методички Минтруда", "UK «Особенность ОТ»", "Правильная ссылка для записей обучения", "https://www.hse.gov.uk/pubns/indg345.htm"),
    ("training records", "Охрана труда UK", "англ.", "Записи о прохождении обучения (бумажные или электронные).", "Журнал инструктажа / протокол", "UK, США OSHA", "", "https://www.hse.gov.uk/pubns/indg345.htm"),
    ("refresher", "Охрана труда UK", "англ.", "Повторное обучение / инструктаж по срокам.", "Повторный инструктаж", "UK INDG345", "", "https://www.hse.gov.uk/pubns/indg345.htm"),
    ("risk assessment (RA)", "Охрана труда", "англ.", "Оценка рисков на рабочем месте.", "Оценка профессиональных рисков (ОПР)", "UK, ЕС, IE Redmond", "", "https://osha.europa.eu/en/legislation/directives/the-osh-framework-directive/1"),
    # --- Охрана труда США ---
    ("OSHA", "Охрана труда США", "США", "Occupational Safety and Health Administration — федеральная инспекция труда США.", "ГИТ / Ростехнадзор (отраслево)", "США", "", "https://www.osha.gov/"),
    ("LOI", "Охрана труда США", "США", "Letter of Interpretation — официальное разъяснение OSHA по применению норм.", "Письмо Роструда / разъяснение", "США 14.08.1997", "LOI 1997 — electronic training records", "https://www.osha.gov/laws-regs/standardinterpretations/1997-08-14"),
    ("29 CFR Part 1904", "Охрана труда США", "США", "Правила учёта травм и профзаболеваний на производстве.", "Формы учёта НС / статотчётность", "США", "Injury Tracking Application — электронная отчётность", "https://www.osha.gov/laws-regs/regulations/standardnumber/1904"),
    ("ITA", "Охрана труда США", "США", "Injury Tracking Application — онлайн‑система OSHA для отчётности по травмам.", "Электронная отчётность по НС", "США", "", "https://www.osha.gov/injuryreporting/ita"),
    ("electronic training records", "Охрана труда США", "англ.", "Электронные записи об обучении — OSHA 1997 прямо допускает, если их можно предъявить инспектору.", "Электронный протокол / журнал (в РФ журнал запрещён)", "США", "Редкий прямой positive precedent", "https://www.osha.gov/laws-regs/standardinterpretations/1997-08-14"),
    # --- Охрана труда ЕС и др. ---
    ("OSH Framework Directive 89/391", "Охрана труда ЕС", "ЕС", "Рамочная директива об охране труда: обязанности работодателя, документирование, обучение.", "ФЗ об ОТ, ТК", "ЕС", "Не запрещает e‑журнал на уровне Союза", "https://osha.europa.eu/en/legislation/directives/the-osh-framework-directive/1"),
    ("OiRA", "Охрана труда ЕС", "ЕС", "Online interactive Risk Assessment — бесплатные шаблоны оценки рисков EU-OSHA.", "Шаблоны карт ОПР", "ЕС «Что взять»", "", "https://osha.europa.eu/en/tools-and-publications/tools/online-interactive-risk-assessment-oira"),
    ("Unterweisung", "Охрана труда DE", "Германия", "Инструктаж по охране труда (ArbSchG § 12): содержание и связь с Gefährdungsbeurteilung важнее росписи в журнале.", "Инструктаж по ПП 2464", "Германия", "BAG 1 ABR 104/09 — аналогия", "https://www.gesetze-im-internet.de/arbschg/__12.html"),
    ("Gefährdungsbeurteilung", "Охрана труда DE", "Германия", "Оценка опасностей (рисков) на рабочем месте — ядро немецкой системы OSH.", "ОПР", "Германия, матрица документов", "", "https://www.gesetze-im-internet.de/arbschg/__5.html"),
    ("ArbSchG", "Охрана труда DE", "Германия", "Arbeitsschutzgesetz — закон об охране труда.", "ФЗ об ОТ", "Германия", "", "https://www.gesetze-im-internet.de/arbschg/"),
    ("Form 6", "Охрана труда RS", "Сербия", "Форма учёта обученных безопасной работе и СИЗ; с 2025 г. только на бумаге, остальные формы ОТ — электронно с QES.", "Журнал инструктажа", "Сербия", "Точечнее, чем запрет ч. 3 ст. 22.1 ТК", "https://www.paragraf.rs/propisi/zakon-o-bezbednosti-i-zdravlju-na-radu.html"),
    ("LPRL", "Охрана труда ES", "Испания", "Ley de Prevención de Riesgos Laborales — закон о профилактике профрисков.", "ФЗ об ОТ", "Испания", "art. 19 — обучение", "https://www.boe.es/"),
    ("WSH / TRS", "Охрана труда SG", "Сингапур", "Workplace Safety and Health + Training Records System — госреестр обязательного обучения.", "Реестр обученных по 2464", "Сингапур", "MOM Check — проверка до допуска", "https://www.mom.gov.sg/eservices/services/check-worker-training-records"),
    ("MOM", "Охрана труда SG", "Сингапур", "Ministry of Manpower — министерство труда, оператор TRS.", "Минтруд", "Сингапур", "", "https://www.mom.gov.sg/"),
    # --- Наряды, LMS, прочее ---
    ("e-PTW / permit-to-work", "Процессы ОТ", "англ.", "Электронный наряд‑допуск (permit to work) с маршрутом согласований и блокировкой допуска.", "Наряд‑допуск (не в ч. 3 ст. 22.1)", "UK, Беларусь", "", U["by175"][1]),
    ("LMS", "Процессы ОТ", "англ.", "Learning Management System — система дистанционного обучения.", "Корпоративная платформа обучения", "UK, ЕС", "Клик в LMS ≠ журнал инструктажа без реквизитов", ""),
    ("GDPR", "ПДн", "ЕС", "General Data Protection Regulation — регламент о защите персональных данных в ЕС.", "152‑ФЗ", "Сроки и ПДн", "", U["gdpr"][1]),
    # --- Суды и дела (имена) ---
    ("BAG", "Судебная практика", "Германия", "Bundesarbeitsgericht — федеральный трудовой суд.", "ВС РФ / кассация по трудовым", "Германия 1 ABR 104/09", "Аналогия по Unterweisung", "https://www.bundesarbeitsgericht.de/entscheidung/1-abr-104-09/"),
    ("BGH", "Судебная практика", "Германия", "Bundesgerichtshof — федеральный суд общей юрисдикции.", "ВС РФ", "Zugang VIII ZR 155/23", "", "https://www.bundesgerichtshof.de/"),
    ("Redmond 2024", "Судебная практика", "Ирландия", "Дело High Court 2024 IEHC 167 — пробел в оценке рисков (RA), не спор об e‑журнале.", "49‑АД24‑22‑К6 (содержание журнала)", "Ирландия", "Тип «аналогия»", "https://www.bailii.org/ie/cases/IEHC/2024/2024IEHC167.html"),
    ("Cass.", "Судебная практика", "Франция", "Cour de cassation — верховный суд по гражданским делам.", "ВС РФ", "Франция 21‑19.841", "Скан подписи ≠ ЭП", "https://www.legifrance.gouv.fr/"),
    # --- Латиница НИР ---
    ("de lege ferenda", "Формулы НИР", "лат.", "«С точки зрения желаемого права» — предложение изменить закон.", "Проект нормы", "«Что взять», правка 22.1", "", U["tk"][1]),
    ("de lege lata", "Формулы НИР", "лат.", "«По действующему праву» — как есть сейчас.", "Действующая норма", "Сравнение с РФ", "", U["tk"][1]),
    ("анalogия", "Формулы НИР", "рус.", "В таблице: нет опубликованного дела именно об электронном журнале ОТ; карточка про соседний сюжет.", "Не путать с прямой практикой", "Тип карточки", "", U["vs49"][1]),
    # --- Российские аббревиатуры в таблице (для полноты) ---
    ("КЭДО", "РФ (для сравнения)", "рус.", "Кадровый электронный документооборот — ст. 22.1–22.3 ТК.", "—", "РФ, сравнение", "Ч. 3 исключает журнал инструктажа", U["tk"][1]),
    ("ПЭП / УНЭП / УКЭП", "РФ (для сравнения)", "рус.", "Три уровня ЭП по 63‑ФЗ.", "SES / AES / QES", "Вся таблица", "", U["fz63"][1]),
    ("ДТС", "РФ (для сравнения)", "рус.", "Доверенная третья сторона ЕАЭС — проверка иностранной ЭЦП между PKI государств‑членов.", "Ст. 7 63‑ФЗ + Решение ЕЭК № 120", "Трансграничность", "", RF_SOURCES[4][1]),
]

HEADERS = [
    "№",
    "Термин (как в таблице)",
    "Категория",
    "Язык / страна",
    "Расшифровка по‑русски",
    "Ближайший аналог в РФ",
    "Где встречается в сводной таблице",
    "Важно не перепутать",
    "Проверить",
]

CATEGORIES = [
    "Подпись и ЭДО",
    "Международные модели",
    "Доказательства",
    "Охрана труда UK",
    "Охрана труда США",
    "Охрана труда ЕС",
    "Охрана труда DE",
    "Охрана труда RS",
    "Охрана труда ES",
    "Охрана труда SG",
    "Процессы ОТ",
    "ПДн",
    "Судебная практика",
    "Формулы НИР",
    "РФ (для сравнения)",
]


def write_xlsx() -> Path:
    wb = Workbook()
    toc = wb.active
    toc.title = "Содержание"
    toc["A1"] = "Глоссарий иностранных терминов сводной таблицы НИР по ЭДО"
    toc["A1"].font = TITLE
    toc.merge_cells("A1:D1")
    toc["A2"] = (
        "Все непонятные английские, немецкие, латинские слова и аббревиатуры из СВОДНАЯ_ТАБЛИЦА_НИР_ЭДО.xlsx. "
        "Читайте «Расшифровку по‑русски» — это главное. «Проверить» — официальный источник. "
        "Если термин помечен как формула НИР — это не цитата закона."
    )
    toc["A2"].font = Font(name="Calibri", size=10, italic=True, color="535353")
    toc["A2"].alignment = WRAP
    toc.merge_cells("A2:D2")
    toc.row_dimensions[2].height = 48
    toc["A4"] = "Категория"
    toc["B4"] = "Лист"
    toc["C4"] = "Терминов"
    toc["D4"] = "О чём блок"
    for c in range(1, 5):
        toc.cell(4, c).fill = HEADER_FILL
        toc.cell(4, c).font = HEADER_FONT
        toc.cell(4, c).border = THIN
    hints = {
        "Подпись и ЭДО": "SES, AES, QES, eIDAS, ESIGN, ECA, Textform…",
        "Международные модели": "UNCITRAL, MLEC, MLETR…",
        "Доказательства": "audit trail, 106B, LEC, holding…",
        "Охрана труда UK": "HSE, MHSWR, sufficient record, goal-setting…",
        "Охрана труда США": "OSHA, LOI, electronic training records…",
        "Охрана труда ЕС": "89/391, OiRA…",
        "Охрана труда DE": "Unterweisung, Gefährdungsbeurteilung…",
        "Охрана труда RS": "Form 6…",
        "Охрана труда ES": "LPRL, pantallazos…",
        "Охрана труда SG": "TRS, WSH, MOM…",
        "Процессы ОТ": "e-PTW, LMS…",
        "ПДн": "GDPR…",
        "Судебная практика": "BAG, BGH, Redmond…",
        "Формулы НИР": "de lege ferenda, аналогия…",
        "РФ (для сравнения)": "КЭДО, ПЭП, ДТС…",
    }
    row = 5
    for cat in CATEGORIES:
        items = [e for e in ENTRIES if e[1] == cat]
        if not items:
            continue
        sheet = cat[:31].replace("/", " ")
        toc.cell(row, 1, cat).fill = CAT_FILL
        toc.cell(row, 1).border = THIN
        toc.cell(row, 2, sheet).font = LINK
        toc.cell(row, 2).hyperlink = XLHyperlink(
            ref=toc.cell(row, 2).coordinate,
            location=f"{quote_sheetname(sheet)}!A1",
            display=sheet,
        )
        toc.cell(row, 2).border = THIN
        toc.cell(row, 3, len(items)).border = THIN
        toc.cell(row, 4, hints.get(cat, "")).alignment = WRAP
        toc.cell(row, 4).border = THIN
        toc.row_dimensions[row].height = 22
        row += 1
    toc.column_dimensions["A"].width = 28
    toc.column_dimensions["B"].width = 24
    toc.column_dimensions["C"].width = 10
    toc.column_dimensions["D"].width = 52

    # All terms sheet
    all_ws = wb.create_sheet("Все_термины")
    write_sheet(all_ws, ENTRIES, "Все иностранные термины (алфавитный порядок внутри категорий)")

    for cat in CATEGORIES:
        items = [e for e in ENTRIES if e[1] == cat]
        if not items:
            continue
        name = cat[:31].replace("/", " ")
        write_sheet(wb.create_sheet(name), items, f"Категория: {cat}")

    wb.save(OUT_XLSX)
    return OUT_XLSX


def write_sheet(ws, items: list, title: str) -> None:
    ws.merge_cells("A1:I1")
    ws["A1"] = title
    ws["A1"].font = TITLE
    ws.merge_cells("A2:I2")
    ws["A2"] = "Столбец «Расшифровка по‑русски» — главный. «Важно не перепутать» — типичная ошибка при чтении сводной таблицы."
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="535353")
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 28
    hr = 4
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(hr, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN
    for i, e in enumerate(items, 1):
        r = hr + i
        vals = [i, e[0], e[1], e[2], e[3], e[4], e[5], e[6], e[7]]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(r, c, val or "")
            cell.alignment = WRAP
            cell.border = THIN
            cell.font = BODY
            if c == 9 and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.font = LINK
        ws.row_dimensions[r].height = 72
    ws.auto_filter.ref = f"A{hr}:I{hr + len(items)}"
    ws.freeze_panes = "E5"
    widths = [5, 22, 18, 14, 52, 28, 28, 32, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False


def write_docx() -> Path:
    doc = setup_doc()
    add_title_block(
        doc,
        "Глоссарий иностранных терминов",
        "Расшифровка слов и аббревиатур из сводной таблицы НИР по ЭДО (август 2026)",
    )
    add_heading(doc, "Как пользоваться")
    add_para(
        doc,
        "Этот файл дополняет СВОДНАЯ_ТАБЛИЦА_НИР_ЭДО.xlsx. Если в таблице встретили "
        "sufficient record, goal-setting, Unterweisung, QES, audit trail и т.п. — ищите термин здесь "
        "или в Excel‑версии с тем же именем. Жирным в таблице ниже — сам термин; "
        "расшифровка — простым русским языком.",
    )
    add_para(
        doc,
        "Важно: многие английские слова в таблице — не названия статей закона, а сжатые формулы составителя. "
        "Пример: sufficient record — не цитата MHSWR; suitable and sufficient в regulation 3 — про оценку рисков.",
    )
    by_cat: dict[str, list] = {}
    for e in ENTRIES:
        by_cat.setdefault(e[1], []).append(e)
    for cat in CATEGORIES:
        if cat not in by_cat:
            continue
        add_heading(doc, cat)
        rows = []
        for term, _, lang, meaning, analog, where, note, _url in by_cat[cat]:
            rows.append((term, meaning, analog or "—", where))
        add_table(doc, ("Термин", "По‑русски", "Аналог в РФ", "Где в таблице"), rows)
    add_heading(doc, "Источники")
    add_source_list(
        doc,
        [
            U["eidas"],
            U["fz63"],
            U["tk"],
            U["p2464"],
            U["by175"],
            U["gdpr"],
            ("HSE INDG345", "https://www.hse.gov.uk/pubns/indg345.htm"),
            ("OSHA LOI 1997", "https://www.osha.gov/laws-regs/standardinterpretations/1997-08-14"),
            ("ECA 2000 s.7", "https://www.legislation.gov.uk/ukpga/2000/7/section/7"),
        ],
    )
    return save(doc, OUT_DOCX)


def main():
    x = write_xlsx()
    d = write_docx()
    print(f"xlsx={x} size={x.stat().st_size} terms={len(ENTRIES)}")
    print(f"docx={d} size={d.stat().st_size}")


if __name__ == "__main__":
    main()
