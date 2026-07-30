#!/usr/bin/env python3
"""Пересчёт листов аналитики для файла судебной практики."""

import re
from collections import Counter, defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

INPUT_PATH = "/home/ubuntu/.cursor/projects/workspace/uploads/________________________111_3bcb.xlsx"
OUTPUT_PATH = "/workspace/Анализ судебной практики111.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)


def normalize_outcome(value: str) -> str:
    text = str(value or "").strip().lower()
    if "пересмотр" in text:
        return "Пересмотр"
    if "работник" in text and "работодатель" not in text:
        return "Работник"
    if "работодатель" in text:
        return "Работодатель"
    return "Прочее"


def simplify_category(value: str) -> str:
    text = str(value or "").strip().lower()
    if "увольн" in text or "сокращ" in text:
        if "сокращ" in text:
            return "Сокращение"
        return "Увольнение"
    if "дисциплин" in text or "выговор" in text:
        return "Дисциплина / выговор"
    if "гит" in text or "предпис" in text or "административ" in text:
        return "ГИТ / предписания"
    return str(value or "Прочее").strip()


def channel_cluster(value: str) -> str:
    text = str(value or "").strip().lower()
    if not text or "не применял" in text:
        return "Прочее"
    if "гос" in text:
        return "Гос. ЭДО"
    if "e-mail" in text or "email" in text or "почт" in text or "1с" in text:
        return "E-mail / почта"
    if any(k in text for k in ("кэдо", "экдо", "корпоратив", "кадров", "унэп", "нэп", "эп", "эцп", "sms")):
        return "КЭДО / корп. ЭДО"
    return "Прочее"


def legal_cluster(row) -> str:
    group = str(row.get("Группа сравнения") or "").lower()
    channel = str(row.get("Канал / инструмент ЭДО") or "").lower()
    category = str(row.get("Категория процесса") or "").lower()

    if "соглаш" in group or ("соглаш" in channel and "без соглас" not in channel):
        return "КЭДО с соглашением"
    if "нет соглас" in group or "без соглас" in channel or "без кэдо" in channel:
        return "Нет согласия на КЭДО"
    if "sms" in group or "ошибоч" in group or "sms" in channel:
        return "Ошибочное подписание / SMS"
    if "e-mail" in group or "email" in channel or "почт" in channel:
        return "E-mail / почта"
    if "лна" in group or "от" in category or "гит" in category:
        return "Ознакомление ЛНА / ОТ"
    if "дисциплин" in category or "выговор" in category:
        return "Дисциплина через ЭДО"
    cluster = channel_cluster(row.get("Канал / инструмент ЭДО"))
    if cluster == "E-mail / почта":
        return "E-mail / почта"
    if cluster == "КЭДО / корп. ЭДО":
        return "КЭДО с соглашением" if "соглаш" in channel else "Нет согласия на КЭДО"
    return "Прочее"


def pct(part: int, total: int) -> float:
    return round(part / total * 100, 1) if total else 0.0


def dominant_outcome(counts: dict) -> str:
    worker = counts.get("Работник", 0)
    employer = counts.get("Работодатель", 0)
    if worker > employer:
        return "Работник"
    if employer > worker:
        return "Работодатель"
    if worker:
        return "Равно"
    return "—"


def trend_label(worker_pct: float) -> str:
    if worker_pct >= 60:
        return "За работника"
    if worker_pct <= 40:
        return "За работодателя"
    return "Смешанно"


def aggregate_table(df: pd.DataFrame, key_col: str) -> list[dict]:
    rows = []
    for key, part in df.groupby(key_col, dropna=False):
        label = "Без группы" if pd.isna(key) else str(key)
        counts = Counter(part["Исход_норм"])
        total = len(part)
        worker = counts.get("Работник", 0)
        employer = counts.get("Работодатель", 0)
        review = counts.get("Пересмотр", 0)
        rows.append(
            {
                "key": label,
                "total": total,
                "worker": worker,
                "employer": employer,
                "review": review,
                "worker_pct": pct(worker, total),
                "share_pct": pct(total, len(df)),
            }
        )
    rows.sort(key=lambda r: (-r["total"], r["key"]))
    return rows


def set_cell(ws, row, col, value, bold=False, fill=False, align_center=False):
    cell = ws.cell(row=row, column=col, value=value)
    if bold:
        cell.font = HEADER_FONT if fill else Font(bold=True)
    if fill:
        cell.fill = HEADER_FILL
    if align_center:
        cell.alignment = Alignment(horizontal="center")
    return cell


def autosize_columns(ws, max_col=7, min_width=10, max_width=45):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        width = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for value in row:
                if value is not None:
                    width = max(width, min(len(str(value)) + 2, max_width))
        ws.column_dimensions[letter].width = width


def write_table(ws, start_row, headers, rows, col_formats=None):
    for col, header in enumerate(headers, 1):
        set_cell(ws, start_row, col, header, bold=True, fill=True, align_center=True)
    for offset, row in enumerate(rows, 1):
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=start_row + offset, column=col, value=value)
            if col_formats and col in col_formats:
                cell.number_format = col_formats[col]
    return start_row + len(rows) + 2


def build_analytics(df: pd.DataFrame):
    df = df.copy()
    df["Исход_норм"] = df["Исход (на чьей стороне)"].map(normalize_outcome)
    df["Категория_крат"] = df["Категория процесса"].map(simplify_category)
    df["Канал_кластер"] = df["Канал / инструмент ЭДО"].map(channel_cluster)
    df["Правовая_позиция"] = df.apply(legal_cluster, axis=1)

    total = len(df)
    outcome_counts = Counter(df["Исход_норм"])
    worker = outcome_counts.get("Работник", 0)
    employer = outcome_counts.get("Работодатель", 0)
    review = outcome_counts.get("Пересмотр", 0)

    year_rows = aggregate_table(df, "Год")
    year_rows.sort(key=lambda r: r["key"])
    peak_year = max(year_rows, key=lambda r: r["total"])
    category_rows = aggregate_table(df, "Категория_крат")
    channel_rows = aggregate_table(df, "Канал_кластер")
    legal_rows = aggregate_table(df, "Правовая_позиция")

    email_cluster = next((r for r in legal_rows if r["key"] == "E-mail / почта"), None)
    kedo_agree = next((r for r in legal_rows if r["key"] == "КЭДО с соглашением"), None)
    no_consent = next((r for r in legal_rows if r["key"] == "Нет согласия на КЭДО"), None)

    group_rows = []
    grouped = df[df["Группа сравнения"].notna()].groupby("Группа сравнения")
    for key, part in grouped:
        counts = Counter(part["Исход_норм"])
        group_rows.append(
            [
                str(key),
                len(part),
                counts.get("Работник", 0),
                counts.get("Работодатель", 0),
                dominant_outcome(counts),
            ]
        )
    group_rows.sort(key=lambda r: (-r[1], r[0]))

    return {
        "df": df,
        "total": total,
        "worker": worker,
        "employer": employer,
        "review": review,
        "year_rows": year_rows,
        "peak_year": peak_year,
        "category_rows": category_rows,
        "channel_rows": channel_rows,
        "legal_rows": legal_rows,
        "email_cluster": email_cluster,
        "kedo_agree": kedo_agree,
        "no_consent": no_consent,
        "group_rows": group_rows,
    }


def write_analytics_sheet(wb, stats):
    if "Аналитика популярности" in wb.sheetnames:
        del wb["Аналитика популярности"]
    ws = wb.create_sheet("Аналитика популярности", 1)

    total = stats["total"]
    worker = stats["worker"]
    employer = stats["employer"]
    review = stats["review"]
    peak = stats["peak_year"]
    top_cat = stats["category_rows"][0]

    ws["A1"] = "Анализ популярности судебных решений"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = (
        f"«Популярность» = как часто встречается исход внутри года/категории/канала ({total} дел). "
        "Не официальная статистика судов РФ."
    )

    row = 5
    ws.cell(row=row, column=1, value="1. Сводка").font = SECTION_FONT
    row += 1
    summary = [
        ["Показатель", "Значение", "Комментарий"],
        ["Всего дел в выборке", total, "Подборка отобранных дел, не статистика Суддепа"],
        ["Исход в пользу работника", worker, f"{pct(worker, total)}%"],
        ["Исход в пользу работодателя", employer, f"{pct(employer, total)}%"],
        ["Направлено на пересмотр", review, "Кассация без финала"],
        ["Пик по годам", f"{peak['key']} — {peak['total']} дел", f"{peak['share_pct']}% всей выборки"],
        ["Лидер категории", top_cat["key"], f"{top_cat['total']} дел"],
    ]
    email_channel = next((r for r in stats["channel_rows"] if r["key"] == "E-mail / почта"), None)
    if email_channel:
        summary.append(
            [
                "E-mail / почта: за работника",
                f"{email_channel['worker_pct']}%",
                f"{email_channel['total']} дел в выборке",
            ]
        )
    kedo_group = next((r for r in stats["group_rows"] if "КЭДО с соглашением" in r[0]), None)
    if kedo_group:
        total_g, worker_g, employer_g = kedo_group[1], kedo_group[2], kedo_group[3]
        employer_pct = pct(employer_g, total_g)
        summary.append(
            [
                "КЭДО с соглашением: за работодателя",
                f"{employer_pct}%" if employer_pct == 100 else f"{employer_pct}–100%",
                f"{total_g} дел в кластере",
            ]
        )
    elif stats["kedo_agree"]:
        ka = stats["kedo_agree"]
        summary.append(
            [
                "КЭДО с соглашением: за работодателя",
                f"{100 - ka['worker_pct']:.0f}–100%" if ka["worker_pct"] == 0 else f"{ka['worker_pct']}%",
                f"{ka['total']} дел в кластере",
            ]
        )

    for i, line in enumerate(summary):
        for col, value in enumerate(line, 1):
            set_cell(ws, row + i, col, value, bold=(i == 0), fill=(i == 0))

    row += len(summary) + 2
    ws.cell(row=row, column=1, value="2. Динамика по годам").font = SECTION_FONT
    row += 1
    year_table = [
        [
            r["key"],
            r["total"],
            r["worker"],
            r["employer"],
            r["review"],
            r["worker_pct"],
            pct(r["employer"], r["total"]),
        ]
        for r in stats["year_rows"]
    ]
    year_start = row
    row = write_table(
        ws,
        row,
        ["Год", "Всего дел", "Работник", "Работодатель", "Пересмотр", "% в пользу работника", "% в пользу работодателя"],
        year_table,
        col_formats={6: "0.0", 7: "0.0"},
    )

    ws.cell(row=row, column=1, value="3. По категориям").font = SECTION_FONT
    row += 1
    cat_start = row
    cat_table = [
        [r["key"], r["total"], r["worker"], r["employer"], r["worker_pct"], r["share_pct"]]
        for r in stats["category_rows"]
    ]
    row = write_table(
        ws,
        row,
        ["Категория", "Всего", "Работник", "Работодатель", "Доля работника %", "Популярность в выборке %"],
        cat_table,
        col_formats={5: "0.0", 6: "0.0"},
    )

    ws.cell(row=row, column=1, value="4. По каналу ЭДО").font = SECTION_FONT
    row += 1
    ch_start = row
    ch_table = [
        [r["key"], r["total"], r["worker"], r["employer"], r["worker_pct"], r["share_pct"]]
        for r in stats["channel_rows"]
    ]
    row = write_table(
        ws,
        row,
        ["Канал ЭДО", "Всего", "Работник", "Работодатель", "Доля работника %", "Популярность в выборке %"],
        ch_table,
        col_formats={5: "0.0", 6: "0.0"},
    )

    ws.cell(row=row, column=1, value="5. Правовые позиции (кластеры)").font = SECTION_FONT
    row += 1
    legal_start = row
    legal_table = [
        [
            r["key"],
            r["total"],
            r["worker"],
            r["employer"],
            r["worker_pct"],
            trend_label(r["worker_pct"]),
        ]
        for r in stats["legal_rows"]
        if r["key"] != "Прочее"
    ]
    row = write_table(
        ws,
        row,
        ["Правовая позиция", "Дел", "Работник", "Работодатель", "Частота «работник» %", "Тренд"],
        legal_table,
        col_formats={5: "0.0"},
    )

    ws.cell(row=row, column=1, value="6. Группы сравнения").font = SECTION_FONT
    row += 1
    group_start = row
    row = write_table(
        ws,
        row,
        ["Группа сравнения", "Всего", "Работник", "Работодатель", "Преобладающий исход"],
        stats["group_rows"],
    )

    row += 1
    ws.cell(row=row, column=1, value="Выводы:").font = SECTION_FONT
    conclusions = [
        f"• {peak['key']} — пик споров ({peak['total']} дел) после реформы ТК РФ.",
        f"• В выборке {pct(worker, total)}% исходов за работника, {pct(employer, total)}% за работодателя.",
    ]
    if stats["no_consent"]:
        nc = stats["no_consent"]
        conclusions.append(
            f"• «Нет согласия на КЭДО» — устойчивая позиция за работника ({nc['worker_pct']}% в кластере, {nc['total']} дел)."
        )
    for i, text in enumerate(conclusions, 1):
        ws.cell(row=row + i, column=1, value=text)

    autosize_columns(ws)

    # Charts
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Исходы по годам"
    chart1.y_axis.title = "Кол-во дел"
    chart1.x_axis.title = "Год"
    data = Reference(ws, min_col=3, min_row=year_start, max_col=5, max_row=year_start + len(year_table))
    cats = Reference(ws, min_col=1, min_row=year_start + 1, max_row=year_start + len(year_table))
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.width = 18
    chart1.height = 10
    ws.add_chart(chart1, "H3")

    chart2 = LineChart()
    chart2.title = "Доля исходов по годам, %"
    chart2.y_axis.title = "%"
    chart2.x_axis.title = "Год"
    data = Reference(ws, min_col=6, min_row=year_start, max_col=7, max_row=year_start + len(year_table))
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.width = 18
    chart2.height = 10
    ws.add_chart(chart2, "H20")

    chart3 = BarChart()
    chart3.type = "col"
    chart3.title = "По категориям"
    data = Reference(ws, min_col=2, min_row=cat_start, max_col=4, max_row=cat_start + len(cat_table))
    cats = Reference(ws, min_col=1, min_row=cat_start + 1, max_row=cat_start + len(cat_table))
    chart3.add_data(data, titles_from_data=True)
    chart3.set_categories(cats)
    chart3.width = 16
    chart3.height = 10
    ws.add_chart(chart3, "A40")

    chart4 = BarChart()
    chart4.type = "col"
    chart4.title = "По каналу ЭДО"
    data = Reference(ws, min_col=2, min_row=ch_start, max_col=4, max_row=ch_start + len(ch_table))
    cats = Reference(ws, min_col=1, min_row=ch_start + 1, max_row=ch_start + len(ch_table))
    chart4.add_data(data, titles_from_data=True)
    chart4.set_categories(cats)
    chart4.width = 16
    chart4.height = 10
    ws.add_chart(chart4, "H40")

    chart5 = PieChart()
    chart5.title = "Общее распределение исходов"
    pie_row = row - len(stats["group_rows"]) - 4
    ws.cell(row=pie_row, column=8, value="Исход")
    ws.cell(row=pie_row, column=9, value="Кол-во")
    ws.cell(row=pie_row + 1, column=8, value="Работник")
    ws.cell(row=pie_row + 1, column=9, value=worker)
    ws.cell(row=pie_row + 2, column=8, value="Работодатель")
    ws.cell(row=pie_row + 2, column=9, value=employer)
    if review:
        ws.cell(row=pie_row + 3, column=8, value="Пересмотр")
        ws.cell(row=pie_row + 3, column=9, value=review)
        pie_end = pie_row + 3
    else:
        pie_end = pie_row + 2
    data = Reference(ws, min_col=9, min_row=pie_row, max_row=pie_end)
    cats = Reference(ws, min_col=8, min_row=pie_row + 1, max_row=pie_end)
    chart5.add_data(data, titles_from_data=True)
    chart5.set_categories(cats)
    chart5.width = 12
    chart5.height = 10
    ws.add_chart(chart5, "A58")

    chart6 = BarChart()
    chart6.type = "bar"
    chart6.title = "Правовые кластеры"
    data = Reference(ws, min_col=2, min_row=legal_start, max_col=4, max_row=legal_start + len(legal_table))
    cats = Reference(ws, min_col=1, min_row=legal_start + 1, max_row=legal_start + len(legal_table))
    chart6.add_data(data, titles_from_data=True)
    chart6.set_categories(cats)
    chart6.width = 16
    chart6.height = 10
    ws.add_chart(chart6, "H58")


def write_chart_data_sheet(wb, stats):
    if "Данные для графиков" in wb.sheetnames:
        del wb["Данные для графиков"]
    ws = wb.create_sheet("Данные для графиков")

    headers_left = ["Год", "Всего дел", "Работник", "Работодатель", "Пересмотр", "% работник", "% работодатель"]
    for col, header in enumerate(headers_left, 1):
        set_cell(ws, 1, col, header, bold=True, fill=True)

    row = 2
    for item in stats["year_rows"]:
        ws.cell(row=row, column=1, value=item["key"])
        ws.cell(row=row, column=2, value=item["total"])
        ws.cell(row=row, column=3, value=item["worker"])
        ws.cell(row=row, column=4, value=item["employer"])
        ws.cell(row=row, column=5, value=item["review"])
        ws.cell(row=row, column=6, value=item["worker_pct"])
        ws.cell(row=row, column=7, value=pct(item["employer"], item["total"]))
        row += 1

    headers_right = ["Категория", "Всего", "Работник", "Работодатель"]
    for col, header in enumerate(headers_right, 8):
        set_cell(ws, 1, col, header, bold=True, fill=True)
    for i, item in enumerate(stats["category_rows"], 2):
        ws.cell(row=i, column=8, value=item["key"])
        ws.cell(row=i, column=9, value=item["total"])
        ws.cell(row=i, column=10, value=item["worker"])
        ws.cell(row=i, column=11, value=item["employer"])

    outcome_row = row + 2
    set_cell(ws, outcome_row, 1, "Исход", bold=True, fill=True)
    set_cell(ws, outcome_row, 2, "Кол-во", bold=True, fill=True)
    for i, (label, value) in enumerate(
        [
            ("Работник", stats["worker"]),
            ("Работодатель", stats["employer"]),
            ("Пересмотр", stats["review"]),
        ],
        1,
    ):
        ws.cell(row=outcome_row + i, column=1, value=label)
        ws.cell(row=outcome_row + i, column=2, value=value)

    channel_row = outcome_row + 5
    set_cell(ws, channel_row, 1, "Канал ЭДО", bold=True, fill=True)
    set_cell(ws, channel_row, 2, "Всего", bold=True, fill=True)
    set_cell(ws, channel_row, 3, "Работник", bold=True, fill=True)
    set_cell(ws, channel_row, 4, "Работодатель", bold=True, fill=True)
    for i, item in enumerate(stats["channel_rows"], 1):
        ws.cell(row=channel_row + i, column=1, value=item["key"])
        ws.cell(row=channel_row + i, column=2, value=item["total"])
        ws.cell(row=channel_row + i, column=3, value=item["worker"])
        ws.cell(row=channel_row + i, column=4, value=item["employer"])

    autosize_columns(ws, max_col=11)


def write_group_summary_sheet(wb, df):
    if "Сводка по группам" in wb.sheetnames:
        del wb["Сводка по группам"]
    ws = wb.create_sheet("Сводка по группам")

    subset = df[df["Группа сравнения"].notna()].copy()
    subset["Исход_норм"] = subset["Исход (на чьей стороне)"].map(normalize_outcome)
    summary = (
        subset.groupby(["Группа сравнения", "Исход_норм"])
        .size()
        .reset_index(name="Кол-во")
        .sort_values(["Группа сравнения", "Исход_норм"])
    )

    headers = ["Группа сравнения", "Исход (на чьей стороне)", "Кол-во"]
    for col, header in enumerate(headers, 1):
        set_cell(ws, 1, col, header, bold=True, fill=True)
    for i, (_, row) in enumerate(summary.iterrows(), 2):
        ws.cell(row=i, column=1, value=row["Группа сравнения"])
        ws.cell(row=i, column=2, value=row["Исход_норм"])
        ws.cell(row=i, column=3, value=row["Кол-во"])
    autosize_columns(ws, max_col=3)


def main():
    df = pd.read_excel(INPUT_PATH, sheet_name="Судебная практика")
    stats = build_analytics(df)

    wb = load_workbook(INPUT_PATH)
    write_analytics_sheet(wb, stats)
    write_chart_data_sheet(wb, stats)
    write_group_summary_sheet(wb, stats["df"])
    wb.save(OUTPUT_PATH)

    print(f"Saved: {OUTPUT_PATH}")
    print(f"Cases: {stats['total']}")
    print(f"Worker: {stats['worker']}, Employer: {stats['employer']}, Review: {stats['review']}")
    print("Peak year:", stats["peak_year"])


if __name__ == "__main__":
    main()
