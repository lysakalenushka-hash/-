#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Оформление Книга1.xlsx и дополнение пропущенных задач из диалога."""

from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SOURCE = Path("/home/ubuntu/.cursor/projects/workspace/uploads/_____1_4fc1.xlsx")
OUTPUT = Path("Книга1.xlsx")

# Палитра B.7.5 / корпоративный стиль
RED = "E30613"
RED_LIGHT = "FEE2E2"
DARK = "1A1A1A"
GRAY = "6B7280"
GRAY_LIGHT = "F3F4F6"
CREAM = "FEF3C7"
WHITE = "FFFFFF"
BORDER_COLOR = "E5E7EB"

HEADERS = ["Тип", "Название", "Что сделано"]

# Исходные данные + дополнения по диалогу (август, пропущено в таблице)
ROWS = [
    ("month", "Июнь"),
    ("data", "Разработка", "Машинист мостового крана", "Сделана учебная программа, текстовые файлы, презентации"),
    ("data", "Актуализация", "Каталог курсов", "Проверена актуальность ссылок, добавлены недостающие курсы"),
    ("month", "Июль"),
    ("data", "Поиск", "Видео по охране труда", "Найдены каналы и видео с нарушениями по ОТ"),
    ("data", "Анализ", "Судебная практика", "Найдены судебные дела по ЭДО"),
    ("data", "Анализ", "Госключ", "Сравнение использования Госключа разными компаниями"),
    ("data", "Анализ", "Судебная практика", "Найдены судебные дела по обучению по ОТ"),
    ("data", "Анализ", "Наряды-допуски", "Сравнение всех нарядов-допусков"),
    ("data", "Анализ", "Судебная практика", "Найдены судебные дела по НС и инструктажам"),
    ("data", "Поиск", "Видео Уренгой", "Найдены видео для сценария для Газпрома"),
    ("data", "Актуализация", "Вентиляция и тепло", "Актуализированы тесты для РТН"),
    ("data", "Актуализация", "Повторный инструктаж по пожарной безопасности", "Актуализирована презентация и тесты"),
    ("data", "Анализ", "Судебная практика", "Найдены треш-судебные дела по ОТ"),
    ("month", "Август"),
    ("data", "Анализ", "Зарубежная практика", "Проанализирована зарубежная практика для НИРа"),
    ("data", "Актуализация", "Специалист по пожарной профилактике", "Исключена тема 6.3 из презентаций и обр. программы"),
    ("data", "Разработка", "Машинист вышки", "Сделаны презентации"),
    ("data", "Разработка", "Персонал, обслуживающий трубопроводы пара и горячей воды", "Сделаны презентации"),
    ("data", "Разработка", "Г.1.1", "Сделаны презентации"),
    ("data", "Разработка", "Б.8.6.1", "Разработан курс"),
    ("data", "Актуализация", "Первая помощь", "Актуализирован курс по первой помощи"),
    # --- добавлено по диалогу (август, не было в исходной таблице) ---
    ("data", "Разработка", "Б.7.5 — учебный план (газовые сети)", "Учебный план 24 ч (3 модуля), презентации и docx по модулям, zip-архив"),
    ("data", "Разработка", "Б.7.5 — темы 2.1, 2.2, 3.1, 3.2", "Word-материалы для экзамена: минимум ссылок на НПА, максимум текста требований"),
    ("data", "Разработка", "ФНП № 531", "Краткая выжимка ключевых требований (Word)"),
    ("data", "Актуализация", "Общие положения ПБ на ОПО", "Переоформление презентации в красный стиль Б.7.5, все изображения сохранены"),
    ("data", "Разработка", "Б.7.5 — тема 1.2", "Презентация: добавлены пункты 4–6 (материалы идентификации, запрет иных документов, экспертиза → разрешение на строительство)"),
    ("data", "Разработка", "Б.7.5 — тема 2.1", "Презентация «Общие требования к сетям газораспределения и газопотребления» (11 слайдов)"),
]

TYPE_COLORS = {
    "Разработка": "FEE2E2",
    "Актуализация": "DBEAFE",
    "Анализ": "FEF3C7",
    "Поиск": "ECFDF5",
}

thin = Side(style="thin", color=BORDER_COLOR)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws):
    for col, title in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, title)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=RED)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 28


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    style_header(ws)

    row_idx = 2
    data_row = 0
    for item in ROWS:
        kind = item[0]
        if kind == "month":
            month = item[1]
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
            cell = ws.cell(row_idx, 1, month)
            cell.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=DARK)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell.border = BORDER
            ws.row_dimensions[row_idx].height = 26
            row_idx += 1
            continue

        _, typ, name, desc = item
        data_row += 1
        values = [typ, name, desc]
        fill_color = TYPE_COLORS.get(typ, GRAY_LIGHT)
        if data_row % 2 == 0:
            fill_color = "FAFAFA" if typ not in TYPE_COLORS else fill_color

        for col, val in enumerate(values, 1):
            cell = ws.cell(row_idx, col, val)
            cell.font = Font(
                name="Calibri",
                size=10,
                bold=(col == 1),
                color=DARK if col > 1 else RED if col == 1 else DARK,
            )
            if col == 1:
                cell.font = Font(name="Calibri", size=10, bold=True, color=RED)
                cell.fill = PatternFill("solid", fgColor=fill_color)
            else:
                cell.fill = PatternFill("solid", fgColor=fill_color if data_row % 2 else WHITE)
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
                indent=0 if col == 1 else 1,
            )
            cell.border = BORDER
        ws.row_dimensions[row_idx].height = max(36, 14 * (1 + len(desc) // 60))
        row_idx += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 58
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{row_idx - 1}"

    # Лист с пропущенными пунктами (справка)
    ws2 = wb.create_sheet("Пропущено — справка")
    ws2["A1"] = "Что добавлено в таблицу по диалогу (август 2026)"
    ws2["A1"].font = Font(name="Calibri", size=12, bold=True, color=RED)
    notes = [
        "Б.7.5 учебный план 24 ч — был в работе, но не отражён в Книга1",
        "Word по темам 2.1, 2.2, 3.1, 3.2 — экзаменационный формат",
        "ФНП № 531 краткая выжимка — отдельный Word-документ",
        "ОПО — переоформление презентации (красный стиль, картинки сохранены)",
        "Б.7.5 тема 1.2 — 4 новых слайда (пункты 4–6 по идентификации)",
        "Б.7.5 тема 2.1 — презентация из загруженного файла презентация.pptx",
        "",
        "Возможно также не отражено (другие задачи августа, не из этого диалога):",
        "• Сравнение условий ФГОС 20.05.01 и 20.02.04 (Excel + Word)",
        "• Зарубежное законодательство по ЭДО (если отдельно от «Зарубежной практики»)",
    ]
    for i, note in enumerate(notes, 3):
        ws2.cell(i, 1, note).font = Font(name="Calibri", size=10, color=GRAY)
    ws2.column_dimensions["A"].width = 80

    wb.save(OUTPUT)
    print(f"Saved {OUTPUT} ({row_idx - 1} rows)")


if __name__ == "__main__":
    build()
