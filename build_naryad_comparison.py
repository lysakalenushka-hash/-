#!/usr/bin/env python3
"""Сравнение нарядов: структура 1-го файла + дословный текст в зелёных ячейках из 2-го."""

from __future__ import annotations

import re
from pathlib import Path

from docx import Document
from docx.document import Document as Doc
from docx.oxml.ns import qn
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

INPUT = Path("/home/ubuntu/.cursor/projects/workspace/uploads/_____-_______629d.docx")
OUTPUT = Path("/workspace/Сравнение_нарядов-допусков.xlsx")

# Колонки «Группа» + «Составляющая наряда» — как в первом файле.
# Для каждой составляющей: паттерны, которыми достаём ДОСЛОВНЫЙ фрагмент в зелёную ячейку.
FIELDS: list[tuple[str, str, list[str]]] = [
    ("Шапка", "Номер наряда-допуска", [
        r"НАРЯД[-\s]?ДОПУСК\s*[N№n]?\s*[_.…—\-]*",
        r"Наряд-допуск\s*[N№n]?\s*[_.…—\-]*",
    ]),
    ("Шапка", "Организация (наименование)", [
        r"Организация\s*[:_]*",
        r"\(наименование организации\)",
    ]),
    ("Шапка", "Подразделение", [r"Подразделение\s*[:_]*"]),
    ("Шапка", "Место / содержание / характеристика работ", [
        r"поручается",
        r"2\.\s*Наименование работ",
        r"\(наименование работ, место, условия их выполнения\)",
        r"Содержание работ",
        r"Место выполнения работ",
        r"\(содержание, характеристика, место производства и объем работ\)",
        r"необходимые для производства работ",
    ]),
    ("Сроки", "Дата и время начала работ", [
        r"Работу начать:\s*дата",
        r"Начало работ",
        r"1\.3\.\s*Начать работы",
        r"Начать работы",
    ]),
    ("Сроки", "Дата и время окончания работ", [
        r"Работу закончить:\s*дата",
        r"Окончание работ",
        r"1\.4\.\s*Окончить работы",
        r"Окончить работы",
    ]),
    ("Сроки", "Срок действия / «действителен до»", [
        r"Действителен до",
        r"Срок действия наряда",
        r"Выдан\s*[\"«]?",
    ]),
    ("Роли", "Кто выдал наряд (ФИО, должность, подпись)", [
        r"Наряд-допуск выдал",
        r"Наряд выдал",
        r"1\.5\.\s*Наряд выдал",
        r"7\.\s*Наряд-допуск выдал",
        r"Лицо, выдавшее наряд-допуск",
    ]),
    ("Роли", "Производитель работ", [
        r"Производителю\s*работ",
        r"1\.1\.\s*Производителю работ",
        r"Ответственному\s*исполнителю \(производителю\) работ",
        r"Производитель работ",
    ]),
    ("Роли", "Ответственный руководитель работ", [
        r"Ответственному руководителю\s*работ",
        r"Ответственному\s*руководителю работ",
        r"1\.\s*Руководителю работ",
        r"Руководитель работ",
    ]),
    ("Роли", "Допускающий", [
        r"допускающему",
        r"Допускающий к работе",
        r"Допускающий",
    ]),
    ("Роли", "Наблюдающий", [r"наблюдающему", r"наблюдающий"]),
    ("Бригада", "Состав бригады / исполнителей", [
        r"с членами бригады",
        r"Состав исполнителей работ",
        r"с бригадой в составе",
        r"6\.\s*Состав исполнителей работ",
    ]),
    ("Бригада", "ФИО членов бригады", [
        r"Фамилия, имя, отчество \(при наличии\)",
        r"Фамилия, инициалы",
        r"фамилия, инициалы, группа по электробезопасности",
    ]),
    ("Бригада", "Профессия / должность членов бригады", [
        r"Профессия \(должность\), квалификация, группа по электробезопасности",
        r"Профессия \(должность\)",
        r"Профессия",
    ]),
    ("Бригада", "Группа по электробезопасности", [
        r"фамилия,\s*инициалы,\s*группа\s*по\s*электробезопасности",
        r"Профессия \(должность\), квалификация, группа по электробезопасности",
        r"\(фамилия, инициалы, группа по электробезопасности\)",
        r"группа по электробезопасности",
    ]),
    ("Бригада", "Изменения в составе бригады", [
        r"Изменения в составе бригады",
        r"Изменения в составе исполнителей работ",
        r"Введен в состав бригады",
        r"Введен в состав исполнителей работ",
    ]),
    ("Меры", "Мероприятия до начала работ / подготовка рабочих мест", [
        r"Мероприятия по подготовке рабочих мест к выполнению работ",
        r"До\s+начала\s+производства\s+работ\s+необходимо\s+выполнить\s+следующие\s+мероприятия",
        r"До начала работ следует выполнить следующие мероприятия",
        r"При\s+подготовке\s+и\s+производстве\s+работ\s+обеспечить\s+следующие\s+меры\s+безопасности",
        r"1\.2\.\s*При\s+подготовке",
    ]),
    ("Меры", "Мероприятия в процессе работ / меры безопасности", [
        r"В\s+процессе\s+производства\s+работ\s+необходимо\s+выполнить\s+следующие\s+мероприятия",
        r"Наименование мероприятия по безопасности работ на высоте",
    ]),
    ("Меры", "Вредные и опасные производственные факторы (ВОПФ)", [
        r"Вредные\s+и\s+опасные\s+производственные\s+факторы",
        r"Опасные и вредные производственные",
    ]),
    ("Меры", "Отдельные / особые указания", [
        r"Отдельные указания",
        r"Особые условия проведения работ",
    ]),
    ("Меры", "Системы обеспечения безопасности работ на высоте", [
        r"Системы обеспечения безопасности работ на высоте",
    ]),
    ("Инструктаж", "Целевой инструктаж выдающего наряд", [
        r"Регистрация целевого инструктажа, проводимого выдающим наряд-допуск",
        r"Регистрация целевого инструктажа,\s*проводимого выдающим наряд-допуск",
    ]),
    ("Инструктаж", "Целевой инструктаж допускающего", [
        r"Регистрация целевого инструктажа,[\s\n]*проводимого допускающим при первичном допуске",
    ]),
    ("Инструктаж", "Целевой инструктаж руководителя / производителя", [
        r"Регистрация целевого инструктажа,[\s\n]*проводимого ответственным руководителем работ",
        r"Регистрация целевого инструктажа при первичном допуске",
    ]),
    ("Инструктаж", "Инструктаж по ОТ (номера инструкций) + подписи бригады", [
        r"Инструктаж по охране труда в объеме инструкций",
        r"Подпись лица, получившего инструктаж",
        r"\(указать наименования или номера инструкций",
    ]),
    ("Инструктаж", "Подпись об ознакомлении с условиями работ", [
        r"Подпись\s+лица, прошедшего\s+инструктаж\s+и\s+ознакомившегося с условиями",
        r"С условиями работ ознакомлен",
        r"с условиями работ ознакомлен",
        r"С условиями производства работ ознакомлен",
        r"С условиями работы ознакомлен",
    ]),
    ("Допуск", "Разрешение на подготовку РМ / допуск к работам", [
        r"Разрешение на подготовку рабочих мест и на допуск к выполнению работ",
        r"Разрешаю приступить к выполнению работ",
        r"Разрешаю приступить\s*к выполнению работ",
    ]),
    ("Допуск", "Ежедневный допуск / оформление начала и окончания смены", [
        r"Ежедневный допуск к работе и время ее окончания",
        r"Оформление ежедневного допуска к производству работ",
        r"Оформление начала производства работ",
    ]),
    ("Допуск", "Отметка: рабочие места подготовлены / осталось под напряжением", [
        r"Рабочие места подготовлены\.\s*Под напряжением остались",
        r"Рабочие места подготовлены\.",
    ]),
    ("Допуск", "Согласование / разрешение эксплуатирующей организации", [
        r"Письменное\s+разрешение\s+эксплуатирующей\s+организации",
        r"Письменное\s+разрешение\s+\(акт-допуск\)\s+действующего\s+предприятия",
        r"Мероприятия по обеспечению безопасности строительного производства\s+согласованы",
    ]),
    ("Закрытие", "Принятие наряда / получил наряд", [
        r"Наряд-допуск принял",
        r"наряд-допуск получил",
        r"С условиями производства работ ознакомлен, наряд-допуск получил",
        r"С условиями работы ознакомлен, наряд-допуск получил",
        r"С условиями работ ознакомлен и наряд-допуск получил",
    ]),
    ("Закрытие", "Продление наряда-допуска", [
        r"Наряд-допуск продлил",
        r"Наряд продлил",
        r"Наряд-допуск продлен",
        r"Наряд допуск продлен",
    ]),
    ("Закрытие", "Закрытие наряда / окончание работ (подписи)", [
        r"Наряд-допуск закрыт",
        r"Работа полностью закончена, бригада удалена",
        r"Работа\s+выполнена\s+в\s+полном\s+объеме",
        r"Работы\s+завершены,\s+рабочие\s+места\s+убраны",
    ]),
]

ORDER_NAMES = {
    "903н": "Электроустановки",
    "883н": "Строительство",
    "882н": "Дорожные работы",
    "871н": "Автотранспорт",
    "875н": "Горэлектротранспорт",
    "867н": "Объекты связи",
    "866н": "Пищевая продукция",
    "858н": "Водные биоресурсы",
    "835н": "Инструмент",
    "833н": "Технологическое оборудование",
    "832н": "Полиграфия",
    "782н": "Работы на высоте",
    "776н": "Металлопокрытия",
    "814н": "Промтранспорт",
    "758н": "ЖКХ",
    "746н": "Сельское хозяйство",
    "721н": "Метрополитен",
    "644н": "Лес / деревообработка",
    "924н": "Теплоснабжение",
    "922н": "Водолазные",
    "915н": "Нефтепродукты",
    "849н": "Окрасочные работы",
    "834н": "Химвещества / чистка",
    "781н": "Производство цемента",
}

NO = "Нет"
YES_FILL = PatternFill("solid", fgColor="C6EFCE")
NO_FILL = PatternFill("solid", fgColor="F2F2F2")
YES_FONT = Font(color="006100", size=8)
NO_FONT = Font(color="A0A0A0", size=9)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
GROUP_FILLS = {
    "Шапка": PatternFill("solid", fgColor="D6EAF8"),
    "Сроки": PatternFill("solid", fgColor="D5F5E3"),
    "Роли": PatternFill("solid", fgColor="FCF3CF"),
    "Бригада": PatternFill("solid", fgColor="FADBD8"),
    "Меры": PatternFill("solid", fgColor="E8DAEF"),
    "Инструктаж": PatternFill("solid", fgColor="FDEBD0"),
    "Допуск": PatternFill("solid", fgColor="D1F2EB"),
    "Закрытие": PatternFill("solid", fgColor="D5D8DC"),
}
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def norm_space(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()


def iter_block_items(parent):
    parent_elm = parent.element.body if isinstance(parent, Doc) else parent._tc
    for child in parent_elm.iterchildren():
        if child.tag == qn("w:p"):
            yield Paragraph(child, parent)
        elif child.tag == qn("w:tbl"):
            yield Table(child, parent)


def find_literal(text: str, patterns: list[str]) -> str | None:
    for pat in patterns:
        m = re.search(pat, text, flags=re.I | re.DOTALL)
        if m:
            frag = norm_space(m.group(0))
            frag = re.sub(r"[_…—\-]{3,}$", "", frag).strip(" :")
            if len(frag) > 160:
                frag = frag[:157] + "..."
            return frag
    return None


def extract_meta(text: str) -> tuple[str, str, str]:
    compact = norm_space(text)
    m = re.search(r"Правилам по охране труда (.+?),?\s*утвержденн", compact, flags=re.I)
    topic = norm_space(m.group(1).strip(" ,")) if m else ""
    m_order = re.search(r"(?:N|№)\s*([0-9]+н)", compact, flags=re.I)
    order_no = m_order.group(1) if m_order else ""
    m_date = re.search(r"от\s+(\d{1,2}\s+[а-яё]+\s+\d{4}\s*г\.?)", compact, flags=re.I)
    order_date = m_date.group(1) if m_date else ""
    return topic, order_no, order_date


def short_name(title: str, order_no: str) -> str:
    title_l = title.lower()
    if order_no == "922н":
        if "судовых" in title_l:
            return "Водолазные судовые (922н)"
        if "наряд-задание" in title_l:
            return "Водолазные наряд-задание (922н)"
        return "Водолазные работы (922н)"
    base = ORDER_NAMES.get(order_no, title[:30])
    return f"{base} ({order_no})" if order_no else base


def build_forms(doc: Document) -> list[dict]:
    blocks: list[str] = []
    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            t = block.text.strip()
            if t:
                blocks.append(t)
        else:
            cells = []
            for row in block.rows:
                for cell in row.cells:
                    ct = cell.text.strip()
                    if ct:
                        cells.append(ct)
            if cells:
                blocks.append("\n".join(cells))

    starts = [i for i, t in enumerate(blocks) if t.startswith("Приложение")]
    forms = []
    for si, start in enumerate(starts):
        end = starts[si + 1] if si + 1 < len(starts) else len(blocks)
        chunk = blocks[start:end]
        full = "\n".join(chunk)
        appendix = chunk[0]
        topic, order_no, order_date = extract_meta(appendix)
        title = next(
            (norm_space(t) for t in chunk[:12] if re.search(r"НАРЯД|Наряд-допуск|Наряд-задание", t)),
            "Наряд-допуск",
        )
        forms.append(
            {
                "idx": si + 1,
                "name": short_name(title, order_no),
                "title": title[:140],
                "topic": topic,
                "order_no": order_no,
                "order_date": order_date,
                "full": full,
            }
        )

    counts: dict[str, int] = {}
    for f in forms:
        counts[f["name"]] = counts.get(f["name"], 0) + 1
    seen: dict[str, int] = {}
    for f in forms:
        if counts[f["name"]] > 1:
            seen[f["name"]] = seen.get(f["name"], 0) + 1
            f["name"] = f'{f["name"]} #{seen[f["name"]]}'
    return forms


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = THIN


def is_present(val) -> bool:
    return bool(val) and val != NO


def main():
    forms = build_forms(Document(str(INPUT)))

    # matrix: field -> form -> literal text or NO
    matrix: list[dict] = []
    for group, field, patterns in FIELDS:
        row = {"group": group, "field": field}
        for f in forms:
            lit = find_literal(f["full"], patterns)
            row[f["name"]] = lit if lit else NO
        matrix.append(row)

    for f in forms:
        f["need_count"] = sum(1 for r in matrix if is_present(r[f["name"]]))

    wb = Workbook()

    # ========== 1. Матрица (структура 1-го файла, зелёные ячейки — дословно) ==========
    ws = wb.active
    ws.title = "Что указывать (матрица)"
    headers = ["Группа", "Составляющая наряда"] + [f["name"] for f in forms] + ["В скольких нарядах нужно"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(1, col, h))
    ws.row_dimensions[1].height = 70
    ws.freeze_panes = "C2"

    for ri, row in enumerate(matrix, 2):
        g = ws.cell(ri, 1, row["group"])
        g.fill = GROUP_FILLS.get(row["group"], PatternFill())
        g.alignment = Alignment(vertical="center", horizontal="center")
        g.border = THIN
        g.font = Font(bold=True, size=9)

        fcell = ws.cell(ri, 2, row["field"])
        fcell.alignment = Alignment(wrap_text=True, vertical="center")
        fcell.border = THIN
        fcell.fill = GROUP_FILLS.get(row["group"], PatternFill())

        need_n = 0
        for ci, form in enumerate(forms, 3):
            val = row[form["name"]]
            c = ws.cell(ri, ci, val)
            c.border = THIN
            if is_present(val):
                c.fill = YES_FILL
                c.font = YES_FONT
                c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                need_n += 1
            else:
                c.fill = NO_FILL
                c.font = NO_FONT
                c.alignment = Alignment(horizontal="center", vertical="center")
        total = ws.cell(ri, len(forms) + 3, need_n)
        total.alignment = Alignment(horizontal="center", vertical="center")
        total.font = Font(bold=True)
        total.border = THIN
        ws.row_dimensions[ri].height = 40

    tr = len(matrix) + 2
    ws.cell(tr, 1, "").border = THIN
    ws.cell(tr, 2, "Сколько составляющих нужно заполнять").font = Font(bold=True)
    ws.cell(tr, 2).border = THIN
    for ci, form in enumerate(forms, 3):
        c = ws.cell(tr, ci, form["need_count"])
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill("solid", fgColor="DDEBF7")
        c.border = THIN

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 48
    for i in range(3, len(forms) + 4):
        ws.column_dimensions[get_column_letter(i)].width = 18

    # ========== 2. По составляющей ==========
    ws2 = wb.create_sheet("По составляющей")
    for col, h in enumerate(
        ["Группа", "Составляющая", "Нужно указывать (наряды)", "Не нужно (наряды)", "Нужно, шт.", "Нет, шт."],
        1,
    ):
        style_header(ws2.cell(1, col, h))
    ws2.freeze_panes = "A2"

    for ri, row in enumerate(matrix, 2):
        need = [f["name"] for f in forms if is_present(row[f["name"]])]
        skip = [f["name"] for f in forms if not is_present(row[f["name"]])]
        values = [
            row["group"],
            row["field"],
            "; ".join(need) if need else "—",
            "; ".join(skip) if skip else "—",
            len(need),
            len(skip),
        ]
        for ci, v in enumerate(values, 1):
            c = ws2.cell(ri, ci, v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = THIN
            if ci == 1:
                c.fill = GROUP_FILLS.get(row["group"], PatternFill())
                c.font = Font(bold=True, size=9)
            if ci == 3 and need:
                c.fill = YES_FILL
            if ci == 4 and skip:
                c.fill = NO_FILL
        ws2.row_dimensions[ri].height = max(30, 15 * (1 + max(len(need), len(skip)) // 3))

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 55
    ws2.column_dimensions["D"].width = 55
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 10

    # ========== 3. Чек-лист по наряду ==========
    ws3 = wb.create_sheet("Чек-лист по наряду")
    style_header(ws3.cell(1, 1, "Наряд-допуск"))
    style_header(ws3.cell(1, 2, "Приказ"))
    style_header(ws3.cell(1, 3, "Что НУЖНО указывать"))
    style_header(ws3.cell(1, 4, "Чего НЕТ в бланке"))
    style_header(ws3.cell(1, 5, "Нужно, шт."))
    ws3.freeze_panes = "A2"

    for ri, form in enumerate(forms, 2):
        need_items = []
        skip_items = []
        for r in matrix:
            if is_present(r[form["name"]]):
                # в чек-листе: название составляющей + дословная формулировка
                need_items.append(f"• {r['field']}\n  → {r[form['name']]}")
            else:
                skip_items.append(f"• {r['field']}")
        ws3.cell(ri, 1, form["name"]).alignment = Alignment(wrap_text=True, vertical="top")
        ws3.cell(ri, 2, form["order_no"]).alignment = Alignment(horizontal="center", vertical="top")
        c3 = ws3.cell(ri, 3, "\n".join(need_items))
        c3.alignment = Alignment(wrap_text=True, vertical="top")
        c3.fill = YES_FILL
        c4 = ws3.cell(ri, 4, "\n".join(skip_items))
        c4.alignment = Alignment(wrap_text=True, vertical="top")
        c4.fill = NO_FILL
        ws3.cell(ri, 5, len(need_items)).alignment = Alignment(horizontal="center", vertical="top")
        for ci in range(1, 6):
            ws3.cell(ri, ci).border = THIN
        ws3.row_dimensions[ri].height = min(420, max(80, 14 * max(len(need_items), len(skip_items), 4)))

    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 60
    ws3.column_dimensions["D"].width = 45
    ws3.column_dimensions["E"].width = 12

    # ========== 4. Реестр ==========
    ws4 = wb.create_sheet("Реестр нарядов", 0)
    for col, h in enumerate(
        ["№", "Наряд-допуск", "Область ПОТ", "Приказ", "Дата", "Составляющих нужно указать"],
        1,
    ):
        style_header(ws4.cell(1, col, h))
    for ri, f in enumerate(forms, 2):
        for ci, v in enumerate(
            [f["idx"], f["name"], f["topic"], f["order_no"], f["order_date"], f["need_count"]],
            1,
        ):
            c = ws4.cell(ri, ci, v)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.border = THIN
    ws4.column_dimensions["A"].width = 5
    ws4.column_dimensions["B"].width = 36
    ws4.column_dimensions["C"].width = 50
    ws4.column_dimensions["D"].width = 10
    ws4.column_dimensions["E"].width = 22
    ws4.column_dimensions["F"].width = 18

    # ========== 5. Как читать ==========
    ws5 = wb.create_sheet("Как читать")
    ws5["A1"] = "Как читать таблицу"
    ws5["A1"].font = Font(bold=True, size=14)
    notes = [
        "Колонки «Группа» и «Составляющая наряда» — как в первом файле (единый список для сравнения).",
        "Зелёные ячейки — дословная формулировка из бланка этого наряда (как во втором файле).",
        "Серое «Нет» — в бланке такой составляющей нет.",
        "Листы «Реестр нарядов», «По составляющей», «Чек-лист по наряду» — сохранены как в первом файле.",
        "В чек-листе для нужных пунктов дополнительно показана стрелка «→» с дословным текстом из бланка.",
        "Группа по электробезопасности дословно есть только в 903н, 883н и 746н.",
    ]
    for i, n in enumerate(notes, 3):
        ws5.cell(i, 1, f"• {n}").alignment = Alignment(wrap_text=True)
    ws5.column_dimensions["A"].width = 120

    wb.save(OUTPUT)
    print(f"Saved {OUTPUT}")
    print(f"Forms: {len(forms)}, components: {len(FIELDS)}")

    # sample green cells
    for field_name in ["Группа по электробезопасности", "Дата и время начала работ", "Наблюдающий"]:
        row = next(r for r in matrix if r["field"] == field_name)
        print(f"\n{field_name}:")
        for f in forms:
            if is_present(row[f["name"]]):
                print(f"  {f['name']}: {row[f['name']]}")


if __name__ == "__main__":
    main()
